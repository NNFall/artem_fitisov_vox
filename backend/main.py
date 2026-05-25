import asyncio
import base64
import csv
import hashlib
import hmac
import io
import json
import logging
import os
import re
import time
from contextlib import asynccontextmanager
from datetime import datetime, timedelta
from html import escape
from pathlib import Path
from typing import Any, Optional
from urllib.parse import urlparse

import aiohttp
from aiogram import Bot, Dispatcher, F
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.filters import Command
from aiogram.types import BotCommand, FSInputFile, Message
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from dotenv import load_dotenv
from fastapi import Cookie, Depends, FastAPI, File, HTTPException, Path as ApiPath, Query, Request, Response, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field
from sqlalchemy import or_
from sqlalchemy.orm import Session

from database import Call, Campaign, OutboundContact, OutboundEvent, OutboundTask, SessionLocal
from voximplant.apiclient import VoximplantAPI

BASE_DIR = Path(__file__).resolve().parent
LOG_PATH = BASE_DIR / "backend.log"
RECORDINGS_DIR = BASE_DIR / "recordings"
IMPORTS_DIR = BASE_DIR / "imports"
TEMPLATES_DIR = BASE_DIR / "templates"
FORUM_AMIX_TEMPLATE_PATH = TEMPLATES_DIR / "forum_amix_template.xlsx"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(name)s %(message)s",
    handlers=[logging.FileHandler(LOG_PATH, encoding="utf-8"), logging.StreamHandler()],
)
logger = logging.getLogger("artem_fitisov_backend")

load_dotenv(BASE_DIR / ".env")

PROJECT_NAME = "artem_fitisov"
BACKEND_WEBHOOK_SECRET = os.getenv("BACKEND_WEBHOOK_SECRET", "").strip()
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "").strip()
TELEGRAM_ADMIN_CHAT_ID = os.getenv("TELEGRAM_ADMIN_CHAT_ID", "").strip()
TELEGRAM_ADMIN_IDS_STR = os.getenv("TELEGRAM_ADMIN_IDS", "")
TELEGRAM_USER_CHAT_IDS_STR = os.getenv("TELEGRAM_USER_CHAT_IDS", "")
GOOGLE_APPS_SCRIPT_WEBHOOK_URL = os.getenv("GOOGLE_APPS_SCRIPT_WEBHOOK_URL", "").strip()
PUBLIC_BACKEND_URL = os.getenv("PUBLIC_BACKEND_URL", "").strip().rstrip("/")
VOXIMPLANT_RULE_ID = os.getenv("VOXIMPLANT_RULE_ID", "").strip()
OUTBOUND_WORKER_ENABLED = os.getenv("OUTBOUND_WORKER_ENABLED", "true").lower() not in {"0", "false", "no", "off"}
OUTBOUND_WORKER_INTERVAL_SECONDS = max(5, int(os.getenv("OUTBOUND_WORKER_INTERVAL_SECONDS", "15")))
OUTBOUND_MAX_CONCURRENT_CALLS = max(1, int(os.getenv("OUTBOUND_MAX_CONCURRENT_CALLS", "1")))
OUTBOUND_RETRY_DELAY_MINUTES = max(1, int(os.getenv("OUTBOUND_RETRY_DELAY_MINUTES", "30")))
OUTBOUND_DEFAULT_CALL_DELAY_SECONDS = max(0, int(os.getenv("OUTBOUND_DEFAULT_CALL_DELAY_SECONDS", "60")))
RECORDINGS_TTL_DAYS = int(os.getenv("RECORDINGS_TTL_DAYS", "30"))
CLEANUP_INTERVAL_HOURS = int(os.getenv("CLEANUP_INTERVAL_HOURS", "24"))
DELIVERY_RETRY_ATTEMPTS = max(1, int(os.getenv("DELIVERY_RETRY_ATTEMPTS", "3")))
DELIVERY_RETRY_DELAY_MS = max(100, int(os.getenv("DELIVERY_RETRY_DELAY_MS", "700")))
DOWNLOAD_RETRY_ATTEMPTS = max(1, int(os.getenv("DOWNLOAD_RETRY_ATTEMPTS", "3")))
DOWNLOAD_RETRY_DELAY_MS = max(100, int(os.getenv("DOWNLOAD_RETRY_DELAY_MS", "1200")))
RECORDING_DOWNLOAD_RETRY_DELAYS_SECONDS_STR = os.getenv("RECORDING_DOWNLOAD_RETRY_DELAYS_SECONDS", "0,30,60,180")
VOXIMPLANT_CREDENTIALS_FILE_PATH = os.getenv("VOXIMPLANT_CREDENTIALS_FILE_PATH", "").strip()
WEB_ADMIN_USERNAME = os.getenv("WEB_ADMIN_USERNAME", "admin").strip()
WEB_ADMIN_PASSWORD = os.getenv("WEB_ADMIN_PASSWORD", "").strip()
WEB_SESSION_SECRET = (
    os.getenv("WEB_SESSION_SECRET", "").strip()
    or BACKEND_WEBHOOK_SECRET
    or "change-this-web-session-secret"
)
WEB_SESSION_COOKIE_NAME = os.getenv("WEB_SESSION_COOKIE_NAME", "artem_web_session").strip() or "artem_web_session"
WEB_SESSION_TTL_SECONDS = max(3600, int(os.getenv("WEB_SESSION_TTL_SECONDS", "28800")))
WEB_CORS_ORIGINS = [
    item.strip().rstrip("/")
    for item in os.getenv(
        "WEB_CORS_ORIGINS",
        "http://localhost:5173,http://127.0.0.1:5173,http://186.246.18.100:3002,http://186.246.18.100:8082,https://obzvonai.ru,https://www.obzvonai.ru",
    ).split(",")
    if item.strip()
]

RECORDINGS_DIR.mkdir(parents=True, exist_ok=True)
IMPORTS_DIR.mkdir(parents=True, exist_ok=True)
TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)

bot: Optional[Bot] = None
if TELEGRAM_BOT_TOKEN:
    bot = Bot(
        token=TELEGRAM_BOT_TOKEN,
        default=DefaultBotProperties(parse_mode=ParseMode.HTML),
    )
else:
    logger.warning("TELEGRAM_BOT_TOKEN is not configured. Telegram delivery is disabled.")

scheduler = AsyncIOScheduler()
dispatcher: Optional[Dispatcher] = None
telegram_polling_task: Optional[asyncio.Task] = None
voximplant_api_client: Optional[VoximplantAPI] = None

TELEGRAM_BOT_COMMANDS = [
    BotCommand(command="help", description="показать справку"),
    BotCommand(command="upload", description="получить шаблон и инструкцию загрузки"),
    BotCommand(command="template", description="получить XLSX-шаблон базы"),
    BotCommand(command="campaigns", description="список кампаний"),
    BotCommand(command="run", description="запустить кампанию: /run ID"),
    BotCommand(command="rerun", description="подготовить повторный запуск: /rerun ID"),
    BotCommand(command="pause", description="поставить кампанию на паузу: /pause ID"),
    BotCommand(command="stop", description="остановить кампанию: /stop ID"),
    BotCommand(command="delay", description="пауза между контактами: /delay ID 30s"),
    BotCommand(command="status", description="статус сервера и очереди"),
    BotCommand(command="calls", description="последние звонки"),
    BotCommand(command="setcommands", description="обновить меню команд"),
]

if VOXIMPLANT_CREDENTIALS_FILE_PATH:
    try:
        voximplant_api_client = VoximplantAPI(VOXIMPLANT_CREDENTIALS_FILE_PATH)
        logger.info("Voximplant credentials loaded from %s", VOXIMPLANT_CREDENTIALS_FILE_PATH)
    except Exception:  # noqa: BLE001
        logger.exception("Failed to initialize VoximplantAPI from credentials file")


class CallStartedPayload(BaseModel):
    session_id: str
    project: str = PROJECT_NAME
    script_name: Optional[str] = None
    outbound_task_id: Optional[int] = None
    campaign_id: Optional[int] = None
    caller_phone: Optional[str] = None
    connected_at_utc: Optional[str] = None


class CallFinishedPayload(BaseModel):
    session_id: str
    duration: Optional[int] = None
    summary: Optional[str] = None
    finished_at_utc: Optional[str] = None
    status: Optional[str] = None


class RecordingReadyPayload(BaseModel):
    session_id: str
    project: str = PROJECT_NAME
    script_name: Optional[str] = None
    outbound_task_id: Optional[int] = None
    campaign_id: Optional[int] = None
    recording_url: str
    recording_status: Optional[str] = None
    recording_error: Optional[str] = None


class VoximplantStatusPayload(BaseModel):
    stage: str
    status: Optional[str] = None
    message: Optional[str] = None
    session_id: Optional[str] = None
    outbound_task_id: Optional[int] = None
    campaign_id: Optional[int] = None
    phone: Optional[str] = None
    timestamp_utc: Optional[str] = None
    data: dict[str, Any] = Field(default_factory=dict)


class WebLoginPayload(BaseModel):
    username: str
    password: str


class WebDelayPayload(BaseModel):
    delay_seconds: int = Field(ge=0, le=86400)


class WebContactUpdatePayload(BaseModel):
    phone: Optional[str] = None
    name: Optional[str] = None
    company: Optional[str] = None
    city: Optional[str] = None
    source: Optional[str] = None
    task: Optional[str] = None
    context: Optional[str] = None
    preferred_time: Optional[str] = None
    timezone: Optional[str] = None


class CallReportPayload(BaseModel):
    report_type: str
    text: str
    call_id: Optional[str] = None


class FinalizePayload(BaseModel):
    session_id: str
    project: str = PROJECT_NAME
    script_name: Optional[str] = None
    outbound_task_id: Optional[int] = None
    campaign_id: Optional[int] = None
    exported_at_utc: Optional[str] = None
    finalization_reason: Optional[str] = None
    model: Optional[str] = None
    caller_phone: Optional[str] = None
    client_phone: Optional[str] = None
    client_name: Optional[str] = None
    call_duration_sec: Optional[int] = None
    telephony_cost_rub: Optional[float] = None
    websocket_duration_sec: Optional[float] = None
    websocket_cost_rub: Optional[float] = None
    voximplant_total_rub: Optional[float] = None
    ai_cost_usd: Optional[float] = None
    ai_cost_rub: Optional[float] = None
    total_cost_rub: Optional[float] = None
    summary: Optional[str] = None
    call_goal: Optional[str] = None
    manager_offer: Optional[str] = None
    outcome: Optional[str] = None
    next_step: Optional[str] = None
    dialogue_text: Optional[str] = None
    recording_status: Optional[str] = None
    recording_url: Optional[str] = None
    recording_error: Optional[str] = None
    usage: dict[str, Any] = Field(default_factory=dict)
    summary_fields: dict[str, Any] = Field(default_factory=dict)
    dialogue_items: list[dict[str, Any]] = Field(default_factory=list)
    admin_report_html: Optional[str] = None
    summary_report_html: Optional[str] = None


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def _b64encode_json(payload: dict[str, Any]) -> str:
    raw = json.dumps(payload, separators=(",", ":"), ensure_ascii=False).encode("utf-8")
    return base64.urlsafe_b64encode(raw).decode("ascii").rstrip("=")


def _b64decode_json(value: str) -> dict[str, Any]:
    padded = value + "=" * (-len(value) % 4)
    raw = base64.urlsafe_b64decode(padded.encode("ascii"))
    parsed = json.loads(raw.decode("utf-8"))
    return parsed if isinstance(parsed, dict) else {}


def create_web_session_token(username: str) -> str:
    issued_at = int(time.time())
    payload = _b64encode_json({"u": username, "iat": issued_at, "exp": issued_at + WEB_SESSION_TTL_SECONDS})
    signature = hmac.new(WEB_SESSION_SECRET.encode("utf-8"), payload.encode("ascii"), hashlib.sha256).hexdigest()
    return f"{payload}.{signature}"


def verify_web_session_token(token: str) -> Optional[str]:
    if "." not in safe_text(token):
        return None
    payload, signature = token.rsplit(".", 1)
    expected = hmac.new(WEB_SESSION_SECRET.encode("utf-8"), payload.encode("ascii"), hashlib.sha256).hexdigest()
    if not hmac.compare_digest(signature, expected):
        return None
    try:
        data = _b64decode_json(payload)
    except Exception:  # noqa: BLE001
        return None
    if safe_int(data.get("exp")) is None or safe_int(data.get("exp")) < int(time.time()):
        return None
    username = safe_text(data.get("u")).strip()
    return username or None


def require_web_user(session_token: Optional[str] = Cookie(default=None, alias=WEB_SESSION_COOKIE_NAME)) -> str:
    username = verify_web_session_token(session_token or "")
    if not username:
        raise HTTPException(status_code=401, detail="not authenticated")
    return username


def set_web_session_cookie(response: Response, username: str):
    response.set_cookie(
        key=WEB_SESSION_COOKIE_NAME,
        value=create_web_session_token(username),
        max_age=WEB_SESSION_TTL_SECONDS,
        httponly=True,
        samesite="lax",
        secure=False,
        path="/",
    )


def clear_web_session_cookie(response: Response):
    response.delete_cookie(key=WEB_SESSION_COOKIE_NAME, path="/")


def safe_text(value: Any) -> str:
    return "" if value is None else str(value)


def safe_float(value: Any) -> Optional[float]:
    if value in ("", None):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if number == number else None


def safe_int(value: Any) -> Optional[int]:
    if value in ("", None):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def to_json_text(value: Any) -> Optional[str]:
    if value in (None, "", [], {}):
        return None
    return json.dumps(value, ensure_ascii=False)


def parse_iso_datetime(value: Any) -> Optional[datetime]:
    text = safe_text(value).strip()
    if not text:
        return None

    normalized = text.replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        return None

    if parsed.tzinfo:
        return parsed.astimezone().replace(tzinfo=None)
    return parsed


def normalize_chat_ids(raw_value: str) -> list[str]:
    normalized = safe_text(raw_value).replace(";", ",").replace("\n", ",")
    return [item.strip() for item in normalized.split(",") if item.strip()]


def get_admin_ids() -> set[str]:
    return set(normalize_chat_ids(",".join([TELEGRAM_ADMIN_CHAT_ID, TELEGRAM_ADMIN_IDS_STR])))


def is_admin_chat(chat_id: Any) -> bool:
    admin_ids = get_admin_ids()
    return bool(admin_ids) and safe_text(chat_id) in admin_ids


PHONE_RE = re.compile(r"[^\d+]")


def import_cell_text(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return safe_text(value).strip()


def normalize_phone(value: Any) -> str:
    text = import_cell_text(value)
    candidates = re.findall(r"\+?\d[\d\s\-()]{7,}\d", text)
    phone_source = candidates[0] if candidates else text
    phone = PHONE_RE.sub("", phone_source).strip()
    if phone.startswith("+"):
        phone = phone[1:]
    if phone.endswith("0") and "." in text and text.endswith(".0"):
        phone = phone[:-1]
    if phone.startswith("8") and len(phone) == 11:
        phone = "7" + phone[1:]
    return phone


def normalize_header(value: Any) -> str:
    return safe_text(value).strip().lower().replace("ё", "е")


def dedupe(items: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for item in items:
        if item in seen:
            continue
        seen.add(item)
        result.append(item)
    return result


def backoff_seconds(base_delay_ms: int, attempt: int) -> float:
    return (base_delay_ms * attempt) / 1000.0


async def wait_before_retry(base_delay_ms: int, attempt: int):
    await asyncio.sleep(backoff_seconds(base_delay_ms, attempt))


def recording_download_retry_delays_seconds() -> list[int]:
    delays: list[int] = []
    for item in RECORDING_DOWNLOAD_RETRY_DELAYS_SECONDS_STR.split(","):
        value = safe_int(item.strip())
        if value is not None and value >= 0:
            delays.append(value)
    return delays or [0, 30, 60, 180]


def get_admin_chat_ids() -> list[str]:
    return dedupe(normalize_chat_ids(",".join([TELEGRAM_ADMIN_CHAT_ID, TELEGRAM_ADMIN_IDS_STR])))


def get_summary_chat_ids() -> list[str]:
    return dedupe(get_admin_chat_ids() + normalize_chat_ids(TELEGRAM_USER_CHAT_IDS_STR))


def require_webhook_secret(request: Request):
    if not BACKEND_WEBHOOK_SECRET:
        return

    supplied = request.headers.get("X-Webhook-Secret", "").strip()
    if supplied != BACKEND_WEBHOOK_SECRET:
        raise HTTPException(status_code=403, detail="invalid webhook secret")


def get_or_create_call(db: Session, session_id: str) -> Call:
    db_call = db.query(Call).filter(Call.voximplant_session_id == session_id).first()
    if db_call:
        return db_call

    db_call = Call(
        voximplant_session_id=session_id,
        status="started",
        started_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
    )
    db.add(db_call)
    db.flush()
    return db_call


def set_if_value(obj: Any, field_name: str, value: Any):
    if value in ("", None):
        return
    setattr(obj, field_name, value)


QUESTION_MARKS_RE = re.compile(r"^[\s?]+$")


def display_or_default(value: Any, default: str = "\u043d\u0435 \u0443\u043a\u0430\u0437\u0430\u043d\u043e") -> str:
    text = safe_text(value).strip()
    if not text:
        return default
    if QUESTION_MARKS_RE.fullmatch(text):
        return default
    return text


def is_diagnostic_finalize(payload: "FinalizePayload") -> bool:
    session_id = safe_text(payload.session_id).strip().lower()
    script_name = safe_text(payload.script_name).strip().lower()
    return session_id.startswith("manual-test-") or script_name in {"manual", "test", "debug"}


DEFAULT_NOT_SPECIFIED = "\u043d\u0435 \u0443\u043a\u0430\u0437\u0430\u043d\u043e"
DEFAULT_UNKNOWN = "\u043d\u0435\u0438\u0437\u0432\u0435\u0441\u0442\u043d\u043e"
DEFAULT_NO_DIALOGUE = "\u0420\u0435\u043f\u043b\u0438\u043a\u0438 \u043d\u0435 \u043d\u0430\u0439\u0434\u0435\u043d\u044b."

AMIX_TEMPLATE_HEADERS = [
    "",
    "Имя",
    "Фамилия",
    "E-mail",
    "Контактныйтелефон",
    "Компания",
    "Пришёл",
    "Дозвон Да/Нет",
    "Вид деятельности",
    "Руководитель компании, Да/Нет",
    "Средний чек",
    "Трафик Сарафан/Входящий",
    "Комментарий",
    "Для бота: Приятно или не приятно говорить с Ботом? Да/Нет",
    "Для бота: Транскрибация диалога",
    "Для бота: Саммари разговора",
]

AMIX_POSITIONAL_PHONE_INDEX = 4
XLSX_HEADER_SCAN_ROWS = 10


def ensure_forum_amix_template() -> Path:
    if FORUM_AMIX_TEMPLATE_PATH.exists():
        return FORUM_AMIX_TEMPLATE_PATH

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(AMIX_TEMPLATE_HEADERS)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    important_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for cell in sheet["G"]:
        cell.fill = important_fill

    widths = {
        "A": 8,
        "B": 16,
        "C": 18,
        "D": 26,
        "E": 20,
        "F": 26,
        "G": 12,
        "H": 16,
        "I": 24,
        "J": 28,
        "K": 16,
        "L": 24,
        "M": 44,
        "N": 36,
        "O": 42,
        "P": 42,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:P1"
    sheet.row_dimensions[1].height = 42
    FORUM_AMIX_TEMPLATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(FORUM_AMIX_TEMPLATE_PATH)
    return FORUM_AMIX_TEMPLATE_PATH


COLUMN_ALIASES = {
    "phone": {"phone", "tel", "telephone", "номер", "телефон", "номер телефона", "phone_number", "контактный телефон", "контактныйтелефон"},
    "name": {"name", "client_name", "имя", "клиент", "фио", "контакт"},
    "last_name": {"last_name", "surname", "фамилия"},
    "email": {"email", "e-mail", "mail", "почта", "электронная почта"},
    "company": {"company", "компания", "организация", "бизнес"},
    "city": {"city", "город"},
    "source": {"source", "источник", "мероприятие", "event"},
    "attendance_status": {"пришел", "пришёл", "посетил", "статус участия"},
    "dial_status": {"дозвон", "дозвон да/нет", "дозвон да нет"},
    "activity_type": {"вид деятельности", "деятельность", "ниша"},
    "is_decision_maker": {"руководитель компании, да/нет", "руководитель компании", "руководитель", "лпр"},
    "average_check": {"средний чек", "чек"},
    "traffic_source": {"трафик сарафан/входящий", "трафик", "сарафан/входящий", "канал клиентов"},
    "bot_impression": {"для бота: приятно или не приятно говорить с ботом? да/нет", "приятно говорить с ботом", "реакция на бота"},
    "dialogue_transcript": {"для бота: транскрибация диалога", "транскрибация диалога", "транскрипт"},
    "call_summary": {"для бота: саммари разговора", "саммари разговора", "summary"},
    "task": {"task", "задача", "интерес", "need", "потребность"},
    "context": {"context", "prompt_context", "контекст", "комментарий", "comment", "info", "доп информация"},
    "preferred_time": {"preferred_time", "удобное время", "time", "время"},
    "timezone": {"timezone", "tz", "часовой пояс"},
    "call_after": {"call_after", "scheduled_at", "когда звонить", "дата звонка"},
    "max_attempts": {"max_attempts", "попытки", "количество попыток"},
    "campaign_context": {"campaign_context", "общий контекст", "общая информация", "global_context"},
}


def canonical_column(header: Any) -> str:
    normalized = normalize_header(header)
    for canonical, aliases in COLUMN_ALIASES.items():
        if normalized in aliases:
            return canonical
    return normalized


def parse_datetime_maybe(value: Any) -> Optional[datetime]:
    text = safe_text(value).strip()
    if not text:
        return None
    parsed = parse_iso_datetime(text)
    if parsed:
        return parsed
    for fmt in ("%d.%m.%Y %H:%M", "%d.%m.%Y", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def read_csv_rows(content: bytes, suffix: str) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    if suffix == ".tsv":
        delimiter = "\t"
    else:
        try:
            delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t").delimiter
        except csv.Error:
            delimiter = ";"
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    return [dict(row) for row in reader if any(safe_text(v).strip() for v in row.values())]


def read_xlsx_rows(content: bytes) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    first_data_row_index = 0
    headers: list[str] = []
    for index, row in enumerate(rows[:XLSX_HEADER_SCAN_ROWS]):
        candidate_headers = [safe_text(cell).strip() for cell in row]
        canonical_headers = {canonical_column(header) for header in candidate_headers if safe_text(header).strip()}
        if "phone" in canonical_headers:
            headers = candidate_headers
            first_data_row_index = index + 1
            break

    if not headers:
        first_non_empty_index = next(
            (index for index, row in enumerate(rows) if any(safe_text(value).strip() for value in row)),
            None,
        )
        if first_non_empty_index is None:
            return []
        first_row = rows[first_non_empty_index]
        if len(first_row) > AMIX_POSITIONAL_PHONE_INDEX and normalize_phone(first_row[AMIX_POSITIONAL_PHONE_INDEX]):
            headers = AMIX_TEMPLATE_HEADERS
            first_data_row_index = first_non_empty_index
        else:
            headers = [safe_text(cell).strip() for cell in rows[0]]
            first_data_row_index = 1

    result: list[dict[str, Any]] = []
    for row in rows[first_data_row_index:]:
        item = {headers[i]: row[i] if i < len(row) else "" for i in range(len(headers)) if headers[i]}
        if any(safe_text(v).strip() for v in item.values()):
            result.append(item)
    return result


def normalize_attendance_status(value: Any) -> str:
    text = normalize_header(value).replace(" ", "")
    if not text:
        return "not_attended"
    if text in {"да", "yes", "y", "1", "пришел", "пришёл", "посетил"}:
        return "attended"
    if text in {"нет", "no", "n", "0", "непришел", "непришёл", "непосетил"}:
        return "not_attended"
    return safe_text(value).strip()


def attendance_label(value: Any) -> str:
    status = safe_text(value).strip()
    if status == "attended":
        return "пришел на мебельный форум Amix"
    if status == "not_attended":
        return "не пришел на мебельный форум Amix"
    return status


def compact_context_lines(lines: list[str]) -> str:
    return "\n".join(line for line in lines if safe_text(line).strip())


def normalize_import_rows(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], str]:
    normalized_rows: list[dict[str, Any]] = []
    campaign_context = ""

    for raw in rows:
        item: dict[str, Any] = {}
        raw_clean = {safe_text(k).strip(): import_cell_text(v) for k, v in raw.items() if safe_text(k).strip()}
        for key, value in raw_clean.items():
            item[canonical_column(key)] = import_cell_text(value)

        first_name = safe_text(item.get("name")).strip()
        last_name = safe_text(item.get("last_name")).strip()
        if last_name and first_name and last_name.lower() not in first_name.lower():
            item["name"] = f"{first_name} {last_name}"
        elif last_name and not first_name:
            item["name"] = last_name

        phone = normalize_phone(item.get("phone"))
        if not phone:
            continue
        item["phone"] = phone

        if "attendance_status" in item:
            item["attendance_status"] = normalize_attendance_status(item.get("attendance_status"))
            item["source"] = item.get("source") or "Форум Amix"

        context_lines: list[str] = []
        if item.get("attendance_status"):
            context_lines.append(f"Статус участия: {attendance_label(item.get('attendance_status'))}.")
        if item.get("activity_type"):
            context_lines.append(f"Вид деятельности: {item['activity_type']}.")
        if item.get("is_decision_maker"):
            context_lines.append(f"Руководитель компании: {item['is_decision_maker']}.")
        if item.get("average_check"):
            context_lines.append(f"Средний чек: {item['average_check']}.")
        if item.get("traffic_source"):
            context_lines.append(f"Канал клиентов: {item['traffic_source']}.")
        if item.get("bot_impression"):
            context_lines.append(f"Оценка разговора с ботом: {item['bot_impression']}.")
        if item.get("context"):
            context_lines.append(item["context"])
        if context_lines:
            item["context"] = compact_context_lines(context_lines)

        if item.get("campaign_context") and not campaign_context:
            campaign_context = safe_text(item.get("campaign_context")).strip()

        item["_raw"] = {
            "source_row": raw_clean,
            "normalized": {k: v for k, v in item.items() if not k.startswith("_")},
        }
        normalized_rows.append(item)

    return normalized_rows, campaign_context


def json_dict(value: Any) -> dict[str, Any]:
    if not value:
        return {}
    if isinstance(value, dict):
        return value
    try:
        parsed = json.loads(safe_text(value))
    except (TypeError, ValueError, json.JSONDecodeError):
        return {}
    return parsed if isinstance(parsed, dict) else {}


def build_task_context(campaign: Campaign, contact: OutboundContact, task: OutboundTask) -> dict[str, Any]:
    raw_row = json_dict(contact.raw_row_json)
    normalized = raw_row.get("normalized") if isinstance(raw_row.get("normalized"), dict) else {}
    return {
        "task_id": task.id,
        "campaign_id": campaign.id,
        "phone": contact.phone,
        "client_name": contact.name or "",
        "last_name": normalized.get("last_name", ""),
        "email": normalized.get("email", ""),
        "company": contact.company or "",
        "city": contact.city or "",
        "source": contact.source or "",
        "attendance_status": normalized.get("attendance_status", ""),
        "activity_type": normalized.get("activity_type", ""),
        "is_decision_maker": normalized.get("is_decision_maker", ""),
        "average_check": normalized.get("average_check", ""),
        "traffic_source": normalized.get("traffic_source", ""),
        "bot_impression": normalized.get("bot_impression", ""),
        "task": contact.task or "",
        "context": contact.context or "",
        "preferred_time": contact.preferred_time or "",
        "timezone": contact.timezone or "",
        "campaign_context": campaign.prompt_context or "",
    }


def call_context_brief(db_call: Call) -> dict[str, Any]:
    return {
        "id": db_call.id,
        "session_id": db_call.voximplant_session_id,
        "outbound_task_id": db_call.outbound_task_id,
        "campaign_id": db_call.campaign_id,
        "status": db_call.status,
        "client_name": db_call.client_name or "",
        "client_phone": db_call.client_phone or "",
        "caller_phone": db_call.caller_phone or "",
        "summary": db_call.summary or "",
        "outcome": db_call.outcome or "",
        "next_step": db_call.next_step or "",
        "call_goal": db_call.call_goal or "",
        "manager_offer": db_call.manager_offer or "",
        "recording_status": db_call.recording_status or "",
        "recording_url": db_call.recording_url or "",
        "started_at": db_call.started_at.isoformat() if db_call.started_at else None,
        "connected_at": db_call.connected_at.isoformat() if db_call.connected_at else None,
        "finished_at": db_call.finished_at.isoformat() if db_call.finished_at else None,
        "updated_at": db_call.updated_at.isoformat() if db_call.updated_at else None,
    }


def task_context_brief(task: OutboundTask, contact: Optional[OutboundContact] = None, campaign: Optional[Campaign] = None) -> dict[str, Any]:
    return {
        "id": task.id,
        "campaign_id": task.campaign_id,
        "campaign_name": campaign.name if campaign else "",
        "contact_id": task.contact_id,
        "phone": task.phone,
        "name": contact.name if contact else "",
        "company": contact.company if contact else "",
        "status": task.status,
        "last_status": task.last_status or "",
        "last_status_message": task.last_status_message or "",
        "result_status": task.result_status or "",
        "result_summary": task.result_summary or "",
        "last_error": task.last_error or "",
        "attempt_count": task.attempt_count,
        "max_attempts": task.max_attempts,
        "voximplant_session_id": task.voximplant_session_id or "",
        "scheduled_at": task.scheduled_at.isoformat() if task.scheduled_at else None,
        "started_at": task.started_at.isoformat() if task.started_at else None,
        "finished_at": task.finished_at.isoformat() if task.finished_at else None,
        "updated_at": task.updated_at.isoformat() if task.updated_at else None,
    }


def build_inbound_context_text(
    phone: str,
    lead_context: dict[str, Any],
    last_task: Optional[dict[str, Any]],
    last_call: Optional[dict[str, Any]],
) -> str:
    lines = [f"Входящий звонок с номера {phone}."]
    client_name = safe_text(lead_context.get("client_name")).strip()
    if client_name:
        lines.append(f"Имя клиента в базе: {client_name}.")
    company = safe_text(lead_context.get("company")).strip()
    if company:
        lines.append(f"Компания/ниша: {company}.")
    attendance_status = safe_text(lead_context.get("attendance_status")).strip()
    if attendance_status:
        lines.append(f"Статус участия в Amix: {attendance_label(attendance_status)}.")
    activity_type = safe_text(lead_context.get("activity_type")).strip()
    if activity_type:
        lines.append(f"Вид деятельности: {activity_type}.")
    campaign_context = safe_text(lead_context.get("campaign_context")).strip()
    if campaign_context:
        lines.append(f"Контекст кампании: {campaign_context}.")
    context = safe_text(lead_context.get("context")).strip()
    if context:
        lines.append(f"Контекст строки базы: {context}.")

    if last_task:
        task_status = safe_text(last_task.get("status")).strip()
        last_status = safe_text(last_task.get("last_status") or last_task.get("result_status")).strip()
        if task_status or last_status:
            lines.append(f"Последняя задача обзвона: статус {task_status or 'не указан'}, последний этап {last_status or 'не указан'}.")
        result_summary = safe_text(last_task.get("result_summary")).strip()
        if result_summary:
            lines.append(f"Итог последней задачи: {result_summary}.")

    if last_call:
        call_status = safe_text(last_call.get("status")).strip()
        outcome = safe_text(last_call.get("outcome")).strip()
        summary = safe_text(last_call.get("summary")).strip()
        if call_status or outcome:
            lines.append(f"Последний звонок: статус {call_status or 'не указан'}, итог: {outcome or 'не указан'}.")
        if summary:
            lines.append(f"Summary последнего звонка: {summary}.")

    return "\n".join(lines)


def get_campaign_stats(db: Session, campaign_id: int) -> dict[str, int]:
    rows = db.query(OutboundTask.status, OutboundTask.id).filter(OutboundTask.campaign_id == campaign_id).all()
    stats: dict[str, int] = {}
    for status, _ in rows:
        stats[status or "unknown"] = stats.get(status or "unknown", 0) + 1
    stats["total"] = len(rows)
    return stats


def parse_delay_seconds(value: Any) -> Optional[int]:
    text = safe_text(value).strip().lower().replace(",", ".")
    if not text:
        return None
    match = re.fullmatch(r"(\d+(?:\.\d+)?)(s|sec|сек|секунд|m|min|мин|минут)?", text)
    if not match:
        return None
    number = float(match.group(1))
    unit = match.group(2) or "s"
    if unit in {"m", "min", "мин", "минут"}:
        number *= 60
    return max(0, int(number))


def format_delay(seconds: Any) -> str:
    value = safe_int(seconds) or 0
    if value >= 60 and value % 60 == 0:
        return f"{value // 60} мин"
    return f"{value} сек"


STATUS_RU = {
    "active": "запущена",
    "paused": "на паузе",
    "pending": "ожидает",
    "scheduled": "запланирована",
    "starting": "запускается",
    "started": "сценарий запущен",
    "in_progress": "идет звонок",
    "completed": "завершено",
    "failed": "ошибка",
    "finished": "завершено",
    "finalized": "завершено",
    "started_no_url": "запись началась, ссылка пока не получена",
    "requested_not_confirmed": "запись запрошена",
    "not_started": "не начато",
    "ready": "готово",
    "disabled": "выключено",
    "error": "ошибка",
    "call_timeout": "не взяли трубку",
    "call_failed": "звонок не состоялся",
    "call_disconnected": "звонок завершен",
    "dial_error": "ошибка набора",
    "empty_call_target": "не передан номер",
    "no_gemini_key": "не задан ключ ИИ",
    "gemini_create_error": "ошибка подключения ИИ",
}


def status_ru(value: Any) -> str:
    raw = safe_text(value).strip()
    if not raw:
        return "не указано"
    return STATUS_RU.get(raw, raw)


def display_text(value: Any) -> str:
    text = safe_text(value)
    replacements = {
        "backend": "сервер",
        "Backend": "Сервер",
        "custom_data": "данные запуска",
        "script_custom_data": "данные запуска",
        "summary": "итоги",
        "Summary": "Итоги",
        "AI": "ИИ",
        "Worker": "обработчик",
        "Rule ID": "ID правила",
        "call_timeout": "не взяли трубку",
        "call_failed": "звонок не состоялся",
        "call_disconnected": "звонок завершен",
        "dial_error": "ошибка набора",
        "gemini_create_error": "ошибка подключения ИИ",
        "no_gemini_key": "не задан ключ ИИ",
        "empty_call_target": "не передан номер",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    return text


def html_text(value: Any) -> str:
    return escape(display_text(value))


def format_stat_line(label: str, value: Any) -> str:
    return f"<b>{label}:</b> {html_text(value)}"


def get_campaign_recipients(campaign: Optional[Campaign]) -> list[str]:
    created_by = [campaign.created_by_chat_id] if campaign and campaign.created_by_chat_id else []
    return dedupe(get_admin_chat_ids() + normalize_chat_ids(",".join(created_by)))


def display_contact(contact: Optional[OutboundContact], task: Optional[OutboundTask] = None) -> str:
    if not contact and not task:
        return "контакт не найден"
    name = safe_text(contact.name if contact else "").strip()
    phone = safe_text((contact.phone if contact else None) or (task.phone if task else "")).strip()
    if name and phone:
        return f"{name}, {phone}"
    return name or phone or "контакт не найден"


STATUS_LABELS = {
    "scenario_started": "Сценарий Voximplant запущен",
    "custom_data_loaded": "Данные задачи получены в сценарии",
    "context_fetch_start": "Сценарий запрашивает контекст задачи",
    "context_fetch_done": "Контекст задачи получен",
    "context_fetch_timeout": "Таймаут получения контекста задачи",
    "context_skipped_custom_data": "Контекст взят из custom_data",
    "lead_context_ready": "Контекст лида готов",
    "empty_call_target": "Номер для звонка не найден",
    "dial_start": "Начинаю набор номера",
    "dial_error": "Ошибка набора номера",
    "call_timeout": "Абонент не ответил за таймаут",
    "call_connected": "Абонент взял трубку",
    "recording_requested": "Запись разговора запрошена",
    "recording_ready": "Запись разговора готова",
    "recording_failed": "Ошибка записи разговора",
    "gemini_warmup_start": "Подключаю AI",
    "gemini_ready": "AI подключен",
    "gemini_error": "Ошибка AI",
    "opening_greeting_sent": "Приветствие отправлено",
    "caller_input_enabled": "Микрофон абонента подключен к AI",
    "summary_request": "Запрошена суммаризация",
    "finalize_start": "Финализирую звонок",
    "finalize_sent": "Итоги отправлены на backend",
    "call_disconnected": "Звонок завершен",
    "call_failed": "Звонок не состоялся",
}


def status_label(stage: str, fallback: Optional[str] = None) -> str:
    clean_stage = safe_text(stage).strip()
    return display_text(fallback or STATUS_LABELS.get(clean_stage, clean_stage or "Статус звонка обновлен"))


def task_status_text(
    task: OutboundTask,
    campaign: Campaign,
    contact: Optional[OutboundContact],
    status_text: str,
    call: Optional[Call] = None,
) -> str:
    lines = [
        f"<b>{html_text(status_text)}</b>",
        "",
        format_stat_line("Кампания", f"#{campaign.id} {campaign.name}"),
        format_stat_line("Задача", f"#{task.id}"),
        format_stat_line("Контакт", display_contact(contact, task)),
        format_stat_line("Статус", status_ru(task.status)),
        format_stat_line("Попытка", f"{task.attempt_count or 0}/{task.max_attempts or 1}"),
    ]
    if task.voximplant_session_id:
        lines.append(f"<b>Сессия Voximplant:</b> <code>{escape(task.voximplant_session_id)}</code>")
    if task.last_status:
        lines.append(format_stat_line("Этап", status_label(task.last_status)))
    if task.last_status_message:
        lines.append(format_stat_line("Детали", task.last_status_message))
    if call:
        lines.append("")
        lines.append(format_stat_line("ID звонка", f"#{call.id}"))
        if call.duration is not None:
            lines.append(format_stat_line("Длительность", f"{call.duration} сек"))
        if call.outcome:
            lines.append(format_stat_line("Итог", call.outcome))
        if call.summary:
            lines.append(format_stat_line("Кратко", call.summary))
    if task.last_error:
        lines.append(format_stat_line("Ошибка", status_ru(task.last_error)))
    return "\n".join(lines)


def render_admin_report(payload: FinalizePayload) -> str:
    usage_text = safe_text(payload.usage).strip() or "{}"
    lines = [
        "<b>\u0417\u0432\u043e\u043d\u043e\u043a \u0437\u0430\u0432\u0435\u0440\u0448\u0435\u043d</b>",
        f"<b>\u041d\u043e\u043c\u0435\u0440:</b> {display_or_default(payload.client_phone or payload.caller_phone or DEFAULT_UNKNOWN, DEFAULT_UNKNOWN)}",
        f"<b>\u0414\u043b\u0438\u0442\u0435\u043b\u044c\u043d\u043e\u0441\u0442\u044c:</b> {safe_text(payload.call_duration_sec or 0)} \u0441\u0435\u043a",
        f"<b>\u0422\u0435\u043b\u0435\u0444\u043e\u043d\u0438\u044f:</b> {safe_text(payload.voximplant_total_rub or payload.telephony_cost_rub or 0)} \u0440\u0443\u0431",
        f"<b>ИИ:</b> {safe_text(payload.ai_cost_rub or 0)} \u0440\u0443\u0431 ({safe_text(payload.ai_cost_usd or 0)} USD)",
        f"<b>\u0418\u0442\u043e\u0433\u043e\u0432\u0430\u044f \u0441\u0442\u043e\u0438\u043c\u043e\u0441\u0442\u044c:</b> {safe_text(payload.total_cost_rub or payload.voximplant_total_rub or payload.telephony_cost_rub or 0)} \u0440\u0443\u0431",
    ]

    if payload.recording_url:
        lines.append(f"<b>\u0417\u0430\u043f\u0438\u0441\u044c:</b> {payload.recording_url}")

    lines.extend(
        [
            "",
            "<b>\u0422\u043e\u043a\u0435\u043d\u044b:</b>",
            usage_text,
            "",
            "<b>\u0414\u0438\u0430\u043b\u043e\u0433:</b>",
            display_or_default(payload.dialogue_text or DEFAULT_NO_DIALOGUE, DEFAULT_NO_DIALOGUE),
        ]
    )
    return "\n".join(lines)

def render_summary_report(payload: FinalizePayload) -> str:
    lines = [
        "<b>\u041d\u043e\u0432\u044b\u0439 \u0437\u0432\u043e\u043d\u043e\u043a (\u0441\u0443\u043c\u043c\u0430\u0440\u0438\u0437\u0430\u0446\u0438\u044f)</b>",
        f"<b>\u041d\u043e\u043c\u0435\u0440:</b> {display_or_default(payload.client_phone or payload.caller_phone or DEFAULT_UNKNOWN, DEFAULT_UNKNOWN)}",
        f"<b>\u0418\u043c\u044f:</b> {display_or_default(payload.client_name, DEFAULT_NOT_SPECIFIED)}",
        f"<b>\u0417\u0430\u043f\u0440\u043e\u0441:</b> {display_or_default(payload.call_goal, DEFAULT_NOT_SPECIFIED)}",
        f"<b>\u0427\u0442\u043e \u043f\u0440\u0435\u0434\u043b\u043e\u0436\u0438\u043b\u0438:</b> {display_or_default(payload.manager_offer, DEFAULT_NOT_SPECIFIED)}",
        f"<b>\u0418\u0442\u043e\u0433:</b> {display_or_default(payload.outcome, DEFAULT_NOT_SPECIFIED)}",
        f"<b>\u0421\u043b\u0435\u0434\u0443\u044e\u0449\u0438\u0439 \u0448\u0430\u0433:</b> {display_or_default(payload.next_step, DEFAULT_NOT_SPECIFIED)}",
    ]
    if payload.recording_url:
        lines.append(f"<b>\u0417\u0430\u043f\u0438\u0441\u044c:</b> {payload.recording_url}")
    lines.extend(["", f"<b>\u041a\u0440\u0430\u0442\u043a\u043e:</b> {display_or_default(payload.summary, DEFAULT_NOT_SPECIFIED)}"])
    return "\n".join(lines)


def render_inbound_summary_report(payload: FinalizePayload) -> str:
    total_cost = payload.total_cost_rub
    if total_cost is None:
        total_cost = payload.voximplant_total_rub or payload.telephony_cost_rub

    lines = [
        "<b>Входящий вызов завершен</b>",
        format_stat_line("Тип", "входящий вызов"),
        format_stat_line("Номер", display_or_default(payload.caller_phone or payload.client_phone, DEFAULT_UNKNOWN)),
        format_stat_line("Имя", display_or_default(payload.client_name, DEFAULT_NOT_SPECIFIED)),
        format_stat_line("Длительность", f"{safe_text(payload.call_duration_sec or 0)} сек"),
        format_stat_line("Телефония", f"{safe_text(payload.voximplant_total_rub or payload.telephony_cost_rub or 0)} руб"),
        format_stat_line("AI", f"{safe_text(payload.ai_cost_rub or 0)} руб"),
        format_stat_line("Итоговая стоимость", f"{safe_text(total_cost or 0)} руб"),
        format_stat_line("Итог", display_or_default(payload.outcome, DEFAULT_NOT_SPECIFIED)),
        format_stat_line("Следующий шаг", display_or_default(payload.next_step, DEFAULT_NOT_SPECIFIED)),
        "",
        format_stat_line("Кратко", display_or_default(payload.summary, DEFAULT_NOT_SPECIFIED)),
    ]
    if payload.recording_url:
        lines.append(format_stat_line("Запись", "скачиваю и отправлю аудио ответом на это сообщение"))
    elif payload.recording_status:
        lines.append(format_stat_line("Запись", status_ru(payload.recording_status)))
    return "\n".join(lines)


def telegram_refs_status(chat_ids: list[str], refs: list[tuple[str, int]]) -> tuple[str, Optional[str]]:
    if not chat_ids:
        return "no_recipients", None
    if bot is None:
        return "skipped_no_bot", None
    if not refs:
        return "error", "no_message_sent"

    sent_chat_ids = {safe_text(chat_id) for chat_id, _message_id in refs}
    expected_chat_ids = {safe_text(chat_id) for chat_id in chat_ids}
    if expected_chat_ids.issubset(sent_chat_ids):
        return "sent", None
    return "partial", "not_all_recipients_received_message"


async def send_telegram_text_with_refs(chat_ids: list[str], html_text: str) -> tuple[str, Optional[str], list[dict[str, Any]]]:
    if not html_text:
        return "empty_message", None, []

    sent_messages = await send_telegram_status_message(chat_ids, html_text)
    status, error_text = telegram_refs_status(chat_ids, sent_messages)
    refs = [{"chat_id": chat_id, "message_id": message_id} for chat_id, message_id in sent_messages]
    return status, error_text, refs


async def send_telegram_text(chat_ids: list[str], html_text: str) -> tuple[str, Optional[str]]:
    if not chat_ids:
        return "no_recipients", None
    if not html_text:
        return "empty_message", None
    if bot is None:
        return "skipped_no_bot", None

    errors: list[str] = []
    for chat_id in chat_ids:
        sent = False
        last_error_text = ""
        for attempt in range(1, DELIVERY_RETRY_ATTEMPTS + 1):
            try:
                await bot.send_message(chat_id=chat_id, text=html_text, disable_web_page_preview=True)
                sent = True
                break
            except Exception as exc:  # noqa: BLE001
                last_error_text = str(exc)
                logger.warning(
                    "Telegram send failed for chat=%s attempt=%s/%s: %s",
                    chat_id,
                    attempt,
                    DELIVERY_RETRY_ATTEMPTS,
                    last_error_text,
                )
                if attempt < DELIVERY_RETRY_ATTEMPTS:
                    await wait_before_retry(DELIVERY_RETRY_DELAY_MS, attempt)
        if not sent:
            errors.append(f"{chat_id}: {last_error_text or 'unknown_error'}")

    if errors and len(errors) == len(chat_ids):
        return "error", "; ".join(errors)
    if errors:
        return "partial", "; ".join(errors)
    return "sent", None


async def send_telegram_status_message(chat_ids: list[str], html_text: str) -> list[tuple[str, int]]:
    if not chat_ids or not html_text or bot is None:
        return []

    sent_messages: list[tuple[str, int]] = []
    for chat_id in chat_ids:
        try:
            message = await bot.send_message(
                chat_id=chat_id,
                text=html_text,
                disable_web_page_preview=True,
            )
            sent_messages.append((safe_text(chat_id), message.message_id))
        except Exception as exc:  # noqa: BLE001
            logger.warning("Telegram status send failed for chat=%s: %s", chat_id, str(exc))
    return sent_messages


def parse_task_message_refs(task: OutboundTask) -> list[dict[str, Any]]:
    refs: list[dict[str, Any]] = []
    raw = safe_text(task.telegram_message_refs_json).strip()
    if raw:
        try:
            parsed = json.loads(raw)
        except json.JSONDecodeError:
            parsed = []
        if isinstance(parsed, list):
            for item in parsed:
                if not isinstance(item, dict):
                    continue
                chat_id = safe_text(item.get("chat_id")).strip()
                message_id = safe_int(item.get("message_id"))
                if chat_id and message_id:
                    refs.append({"chat_id": chat_id, "message_id": message_id})

    if task.telegram_chat_id and task.telegram_message_id:
        refs.append({"chat_id": safe_text(task.telegram_chat_id), "message_id": task.telegram_message_id})

    unique: dict[str, dict[str, Any]] = {}
    for ref in refs:
        unique[safe_text(ref["chat_id"])] = ref
    return list(unique.values())


def save_task_message_refs(task: OutboundTask, refs: list[dict[str, Any]]):
    unique: dict[str, dict[str, Any]] = {}
    for ref in refs:
        chat_id = safe_text(ref.get("chat_id")).strip()
        message_id = safe_int(ref.get("message_id"))
        if chat_id and message_id:
            unique[chat_id] = {"chat_id": chat_id, "message_id": message_id}

    clean_refs = list(unique.values())
    task.telegram_message_refs_json = json.dumps(clean_refs, ensure_ascii=False) if clean_refs else None
    if clean_refs:
        task.telegram_chat_id = clean_refs[0]["chat_id"]
        task.telegram_message_id = clean_refs[0]["message_id"]


async def notify_outbound_task(task_id: int, status_text: str) -> tuple[str, Optional[str]]:
    if bot is None:
        return "skipped_no_bot", None

    db = SessionLocal()
    try:
        task = db.query(OutboundTask).filter(OutboundTask.id == task_id).first()
        if not task:
            return "task_not_found", None
        campaign = db.query(Campaign).filter(Campaign.id == task.campaign_id).first()
        if not campaign:
            return "campaign_not_found", None
        contact = db.query(OutboundContact).filter(OutboundContact.id == task.contact_id).first()
        call = None
        if task.voximplant_session_id:
            call = db.query(Call).filter(Call.voximplant_session_id == task.voximplant_session_id).first()

        html_text = task_status_text(task, campaign, contact, status_text, call)
        recipients = get_campaign_recipients(campaign)
        refs = parse_task_message_refs(task)
        kept_refs: list[dict[str, Any]] = []
        failed_chat_ids: set[str] = set()

        for ref in refs:
            try:
                await bot.edit_message_text(
                    chat_id=ref["chat_id"],
                    message_id=ref["message_id"],
                    text=html_text,
                    disable_web_page_preview=True,
                )
                kept_refs.append(ref)
            except Exception as exc:  # noqa: BLE001
                error_text = str(exc)
                if "message is not modified" in error_text.lower():
                    kept_refs.append(ref)
                    continue
                logger.warning(
                    "Telegram status edit failed task_id=%s chat=%s message=%s: %s",
                    task.id,
                    ref["chat_id"],
                    ref["message_id"],
                    error_text,
                )
                failed_chat_ids.add(safe_text(ref["chat_id"]))

        edited_chat_ids = {safe_text(ref["chat_id"]) for ref in kept_refs}
        missing_chat_ids = [
            chat_id for chat_id in recipients
            if chat_id not in edited_chat_ids or chat_id in failed_chat_ids
        ]

        sent_messages = await send_telegram_status_message(missing_chat_ids, html_text)
        sent_refs = [{"chat_id": chat_id, "message_id": message_id} for chat_id, message_id in sent_messages]
        all_refs = kept_refs + sent_refs
        if not all_refs:
            return "send_failed", "no_message_sent"

        save_task_message_refs(task, all_refs)
        task.updated_at = datetime.utcnow()
        db.commit()
        if kept_refs and sent_refs:
            return "edited_and_sent", None
        if kept_refs:
            return "edited", None
        return "sent", None
    finally:
        db.close()


async def send_outbound_recording(task_id: int, local_path: Optional[str]) -> tuple[str, Optional[str]]:
    if bot is None:
        return "skipped_no_bot", None
    if not local_path:
        return "skipped_no_file", None

    file_path = Path(local_path)
    if not file_path.exists() or not file_path.is_file():
        return "file_not_found", local_path

    db = SessionLocal()
    try:
        task = db.query(OutboundTask).filter(OutboundTask.id == task_id).first()
        if not task:
            return "task_not_found", None
        campaign = db.query(Campaign).filter(Campaign.id == task.campaign_id).first()
        call = None
        if task.voximplant_session_id:
            call = db.query(Call).filter(Call.voximplant_session_id == task.voximplant_session_id).first()
        caption_lines = [f"<b>Запись звонка</b>", format_stat_line("Задача", f"#{task.id}")]
        if call:
            caption_lines.append(format_stat_line("ID звонка", f"#{call.id}"))
            if call.duration is not None:
                caption_lines.append(format_stat_line("Длительность", f"{call.duration} сек"))
        caption = "\n".join(caption_lines)

        refs = parse_task_message_refs(task)
        if not refs:
            refs = [{"chat_id": chat_id, "message_id": None} for chat_id in get_campaign_recipients(campaign)]
        if not refs:
            return "task_message_not_found", None

        errors: list[str] = []
        sent_count = 0
        for ref in refs:
            try:
                kwargs: dict[str, Any] = {
                    "chat_id": ref["chat_id"],
                    "audio": FSInputFile(str(file_path)),
                    "caption": caption,
                    "parse_mode": ParseMode.HTML,
                }
                if ref.get("message_id"):
                    kwargs["reply_to_message_id"] = ref["message_id"]
                await bot.send_audio(**kwargs)
                sent_count += 1
            except Exception as exc:  # noqa: BLE001
                errors.append(f"{ref.get('chat_id')}: {str(exc)}")

        if sent_count and errors:
            return "partial", "; ".join(errors)
        if sent_count:
            return "sent", None
        return "error", "; ".join(errors) or "no_audio_sent"
    except Exception as exc:  # noqa: BLE001
        logger.warning("Telegram recording send failed task_id=%s path=%s: %s", task_id, local_path, str(exc))
        return "error", str(exc)
    finally:
        db.close()


async def send_inbound_recording(
    session_id: str,
    local_path: Optional[str],
    reply_refs: Optional[list[dict[str, Any]]] = None,
) -> tuple[str, Optional[str]]:
    if bot is None:
        return "skipped_no_bot", None
    if not local_path:
        return "skipped_no_file", None

    file_path = Path(local_path)
    if not file_path.exists() or not file_path.is_file():
        return "file_not_found", local_path

    db = SessionLocal()
    try:
        db_call = db.query(Call).filter(Call.voximplant_session_id == session_id).first()
        caption_lines = ["<b>Запись входящего звонка</b>", format_stat_line("Сессия", f"<code>{session_id}</code>")]
        if db_call:
            caption_lines.append(format_stat_line("ID звонка", f"#{db_call.id}"))
            caption_lines.append(format_stat_line("Номер", db_call.caller_phone or db_call.client_phone or "не указан"))
            if db_call.duration is not None:
                caption_lines.append(format_stat_line("Длительность", f"{db_call.duration} сек"))
        caption = "\n".join(caption_lines)

        refs = reply_refs or []
        if not refs:
            refs = [{"chat_id": chat_id, "message_id": None} for chat_id in get_summary_chat_ids()]
        if not refs:
            return "no_recipients", None

        errors: list[str] = []
        sent_count = 0
        for ref in refs:
            try:
                kwargs: dict[str, Any] = {
                    "chat_id": ref["chat_id"],
                    "audio": FSInputFile(str(file_path)),
                    "caption": caption,
                    "parse_mode": ParseMode.HTML,
                }
                if ref.get("message_id"):
                    kwargs["reply_to_message_id"] = ref["message_id"]
                await bot.send_audio(**kwargs)
                sent_count += 1
            except Exception as exc:  # noqa: BLE001
                errors.append(f"{ref.get('chat_id')}: {str(exc)}")

        if sent_count and errors:
            return "partial", "; ".join(errors)
        if sent_count:
            return "sent", None
        return "error", "; ".join(errors) or "no_audio_sent"
    except Exception as exc:  # noqa: BLE001
        logger.warning("Telegram inbound recording send failed session=%s path=%s: %s", session_id, local_path, str(exc))
        return "error", str(exc)
    finally:
        db.close()


async def send_to_google_sheets(payload: dict[str, Any]) -> tuple[str, Optional[str]]:
    if not GOOGLE_APPS_SCRIPT_WEBHOOK_URL:
        return "skipped_no_url", None

    last_error = ""
    for attempt in range(1, DELIVERY_RETRY_ATTEMPTS + 1):
        try:
            timeout = aiohttp.ClientTimeout(total=25)
            async with aiohttp.ClientSession(timeout=timeout) as session:
                async with session.post(GOOGLE_APPS_SCRIPT_WEBHOOK_URL, json=payload) as response:
                    response_text = await response.text()
                    if response.status >= 400:
                        last_error = f"HTTP {response.status}: {response_text[:1000]}"
                        logger.warning(
                            "Google Sheets sync failed attempt=%s/%s: %s",
                            attempt,
                            DELIVERY_RETRY_ATTEMPTS,
                            last_error,
                        )
                    else:
                        return "sent", response_text[:4000]
        except Exception as exc:  # noqa: BLE001
            last_error = str(exc)
            logger.warning(
                "Google Sheets sync network error attempt=%s/%s: %s",
                attempt,
                DELIVERY_RETRY_ATTEMPTS,
                last_error,
            )

        if attempt < DELIVERY_RETRY_ATTEMPTS:
            await wait_before_retry(DELIVERY_RETRY_DELAY_MS, attempt)

    return "error", last_error or "sync_failed"


def guess_recording_extension(url: str) -> str:
    parsed_path = urlparse(url).path.lower()
    for extension in (".mp3", ".wav", ".ogg", ".m4a"):
        if parsed_path.endswith(extension):
            return extension
    return ".mp3"


async def download_recording(url: str, session_id: str) -> tuple[str, Optional[str], Optional[str]]:
    extension = guess_recording_extension(url)
    file_path = RECORDINGS_DIR / f"{session_id}{extension}"
    auth_headers: dict[str, str] = {}

    if voximplant_api_client and "voximplant-records-secure" in safe_text(url):
        try:
            auth_headers["Authorization"] = voximplant_api_client.build_auth_header()
        except Exception as exc:  # noqa: BLE001
            logger.warning(
                "Failed to build authorization header for recording download session=%s: %s",
                session_id,
                str(exc),
            )

    last_error = ""
    retry_delays = recording_download_retry_delays_seconds()
    total_attempts = len(retry_delays)
    for attempt, delay_seconds in enumerate(retry_delays, start=1):
        if delay_seconds > 0:
            logger.info(
                "Waiting %s sec before recording download retry session=%s attempt=%s/%s",
                delay_seconds,
                session_id,
                attempt,
                total_attempts,
            )
            await asyncio.sleep(delay_seconds)
        try:
            timeout = aiohttp.ClientTimeout(total=60)
            async with aiohttp.ClientSession(timeout=timeout) as session:
                async with session.get(url, headers=auth_headers) as response:
                    if response.status >= 400:
                        last_error = f"HTTP {response.status}"
                    else:
                        data = await response.read()
                        if not data:
                            last_error = "empty_recording_file"
                        else:
                            file_path.write_bytes(data)
                            return "downloaded", str(file_path), None
        except Exception as exc:  # noqa: BLE001
            last_error = str(exc)

        logger.warning(
            "Recording download failed session=%s attempt=%s/%s: %s",
            session_id,
            attempt,
            total_attempts,
            last_error,
        )

    return "download_error", None, last_error or "download_failed"


async def persist_recording_download(session_id: str, url: str) -> tuple[str, Optional[str], Optional[str]]:
    status, local_path, error_text = await download_recording(url, session_id)
    db = SessionLocal()
    try:
        db_call = db.query(Call).filter(Call.voximplant_session_id == session_id).first()
        if not db_call:
            return status, local_path, "call_not_found"
        if local_path:
            db_call.local_recording_path = local_path
            db_call.last_error = None
        if error_text:
            db_call.last_error = error_text
        if db_call.recording_status in (None, "", "not_started", "requested_not_confirmed"):
            db_call.recording_status = status
        db_call.updated_at = datetime.utcnow()
        db.commit()
        return status, local_path, error_text
    finally:
        db.close()


async def persist_and_send_outbound_recording(task_id: int, session_id: str, url: str):
    status, local_path, error_text = await persist_recording_download(session_id, url)
    if local_path:
        await notify_outbound_task(task_id, "Запись скачана, отправляю аудио")
        send_status, send_error = await send_outbound_recording(task_id, local_path)
        if send_error:
            logger.warning(
                "Outbound recording send failed task_id=%s status=%s error=%s",
                task_id,
                send_status,
                send_error,
            )
        return

    logger.warning(
        "Outbound recording download failed task_id=%s session=%s status=%s error=%s",
        task_id,
        session_id,
        status,
        error_text,
    )
    await notify_outbound_task(task_id, f"Запись не скачалась: {error_text or status_ru(status)}")


async def persist_and_send_inbound_recording(session_id: str, url: str, reply_refs: Optional[list[dict[str, Any]]] = None):
    status, local_path, error_text = await persist_recording_download(session_id, url)
    if local_path:
        send_status, send_error = await send_inbound_recording(session_id, local_path, reply_refs)
        if send_error:
            logger.warning(
                "Inbound recording send failed session=%s status=%s error=%s",
                session_id,
                send_status,
                send_error,
            )
        return

    logger.warning(
        "Inbound recording download failed session=%s status=%s error=%s",
        session_id,
        status,
        error_text,
    )


async def cleanup_old_recordings():
    logger.info("Starting scheduled cleanup of old recordings")
    cutoff = time.time() - (RECORDINGS_TTL_DAYS * 86400)
    removed_files = 0
    db = SessionLocal()

    try:
        for file_path in RECORDINGS_DIR.iterdir():
            if not file_path.is_file():
                continue
            if file_path.stat().st_mtime >= cutoff:
                continue

            file_path.unlink(missing_ok=True)
            removed_files += 1

            session_id = file_path.stem
            db_call = db.query(Call).filter(Call.voximplant_session_id == session_id).first()
            if db_call:
                db_call.local_recording_path = None
                db_call.updated_at = datetime.utcnow()

        db.commit()
        logger.info("Cleanup finished, removed %s recordings", removed_files)
    except Exception:  # noqa: BLE001
        logger.exception("Cleanup failed")
    finally:
        db.close()


def fill_call_from_finalize(db_call: Call, payload: FinalizePayload):
    now = datetime.utcnow()

    if payload.outbound_task_id:
        db_call.outbound_task_id = payload.outbound_task_id
    if payload.campaign_id:
        db_call.campaign_id = payload.campaign_id

    set_if_value(db_call, "project", payload.project)
    set_if_value(db_call, "script_name", payload.script_name)
    set_if_value(db_call, "model", payload.model)
    set_if_value(db_call, "caller_phone", payload.caller_phone)
    set_if_value(db_call, "client_phone", payload.client_phone)
    set_if_value(db_call, "client_name", payload.client_name)

    exported_at = parse_iso_datetime(payload.exported_at_utc)
    if exported_at:
        db_call.exported_at = exported_at

    if not db_call.connected_at:
        db_call.connected_at = exported_at or now

    db_call.finished_at = now
    db_call.updated_at = now
    db_call.status = payload.finalization_reason or "finalized"

    db_call.duration = safe_int(payload.call_duration_sec)
    db_call.telephony_cost_rub = safe_float(payload.telephony_cost_rub)
    db_call.websocket_duration_sec = safe_float(payload.websocket_duration_sec)
    db_call.websocket_cost_rub = safe_float(payload.websocket_cost_rub)
    db_call.voximplant_total_rub = safe_float(payload.voximplant_total_rub)
    db_call.ai_cost_usd = safe_float(payload.ai_cost_usd)
    db_call.ai_cost_rub = safe_float(payload.ai_cost_rub)
    db_call.total_cost_rub = safe_float(payload.total_cost_rub)

    set_if_value(db_call, "summary", payload.summary)
    set_if_value(db_call, "call_goal", payload.call_goal)
    set_if_value(db_call, "manager_offer", payload.manager_offer)
    set_if_value(db_call, "outcome", payload.outcome)
    set_if_value(db_call, "next_step", payload.next_step)
    set_if_value(db_call, "dialogue_text", payload.dialogue_text)

    set_if_value(db_call, "recording_status", payload.recording_status)
    set_if_value(db_call, "recording_url", payload.recording_url)
    set_if_value(db_call, "recording_error", payload.recording_error)

    db_call.usage_json = to_json_text(payload.usage)
    db_call.summary_fields_json = to_json_text(payload.summary_fields)
    db_call.dialogue_items_json = to_json_text(payload.dialogue_items)
    db_call.admin_report_html = payload.admin_report_html or render_admin_report(payload)
    db_call.summary_report_html = payload.summary_report_html or render_summary_report(payload)
    db_call.raw_payload_json = to_json_text(payload.model_dump(mode="json"))


def update_outbound_task_from_finalize(db: Session, payload: FinalizePayload):
    if not payload.outbound_task_id:
        return

    task = db.query(OutboundTask).filter(OutboundTask.id == payload.outbound_task_id).first()
    if not task:
        return

    now = datetime.utcnow()
    task.voximplant_session_id = payload.session_id
    task.result_status = payload.outcome or payload.finalization_reason or "finalized"
    task.result_summary = payload.summary
    task.updated_at = now
    campaign = db.query(Campaign).filter(Campaign.id == task.campaign_id).first()
    if campaign:
        campaign.next_call_after = now + timedelta(seconds=safe_int(campaign.call_delay_seconds) or 0)
        campaign.updated_at = now

    failed_reasons = {"call_failed", "call_timeout", "no_gemini_key", "gemini_create_error", "empty_call_target"}
    finalization_reason = safe_text(payload.finalization_reason)
    if finalization_reason in failed_reasons and (task.attempt_count or 0) < (task.max_attempts or 1):
        retry_at = now + timedelta(minutes=OUTBOUND_RETRY_DELAY_MINUTES)
        task.status = "pending"
        task.next_attempt_at = retry_at
        task.scheduled_at = retry_at
        task.last_error = finalization_reason
        return

    if finalization_reason in failed_reasons:
        task.status = "failed"
        task.finished_at = now
        task.last_error = finalization_reason
        return

    task.status = "completed"
    task.finished_at = now
    task.last_error = None


async def start_outbound_task(task_id: int):
    if not voximplant_api_client:
        logger.warning("Outbound worker skipped: Voximplant credentials are not configured")
        return
    if not VOXIMPLANT_RULE_ID:
        logger.warning("Outbound worker skipped: VOXIMPLANT_RULE_ID is not configured")
        return

    db = SessionLocal()
    try:
        task = db.query(OutboundTask).filter(OutboundTask.id == task_id).first()
        if not task or task.status not in {"pending", "scheduled"}:
            return
        campaign = db.query(Campaign).filter(Campaign.id == task.campaign_id).first()
        if not campaign or campaign.status != "active":
            return
        contact = db.query(OutboundContact).filter(OutboundContact.id == task.contact_id).first()
        if not contact:
            task.status = "failed"
            task.last_error = "contact_not_found"
            task.updated_at = datetime.utcnow()
            db.commit()
            await notify_outbound_task(task.id, "Контакт не найден")
            return

        now = datetime.utcnow()
        task.status = "starting"
        task.attempt_count = (task.attempt_count or 0) + 1
        task.last_attempt_at = now
        task.updated_at = now
        db.commit()
        await notify_outbound_task(task.id, "Начинаю звонок")

        custom_data = json.dumps(build_task_context(campaign, contact, task), ensure_ascii=False)
        result = await asyncio.to_thread(
            voximplant_api_client.start_scenarios,
            rule_id=int(VOXIMPLANT_RULE_ID),
            script_custom_data=custom_data,
        )

        task.start_result_json = to_json_text(result)
        task.status = "started"
        task.started_at = now
        task.voximplant_session_id = safe_text(result.get("call_session_history_id") or "")
        task.updated_at = datetime.utcnow()
        campaign.next_call_after = datetime.utcnow() + timedelta(seconds=safe_int(campaign.call_delay_seconds) or 0)
        campaign.updated_at = datetime.utcnow()
        db.commit()
        await notify_outbound_task(task.id, "Сценарий Voximplant запущен")
        logger.info("Started outbound task id=%s phone=%s", task.id, task.phone)
    except Exception as exc:  # noqa: BLE001
        logger.exception("Failed to start outbound task id=%s", task_id)
        task = db.query(OutboundTask).filter(OutboundTask.id == task_id).first()
        if task:
            task.last_error = str(exc)
            retry_at = datetime.utcnow() + timedelta(minutes=OUTBOUND_RETRY_DELAY_MINUTES)
            task.status = "pending" if (task.attempt_count or 0) < (task.max_attempts or 1) else "failed"
            task.next_attempt_at = retry_at
            task.scheduled_at = task.next_attempt_at
            task.updated_at = datetime.utcnow()
            db.commit()
            await notify_outbound_task(task.id, "Ошибка запуска звонка")
    finally:
        db.close()


async def process_outbound_queue():
    if not OUTBOUND_WORKER_ENABLED:
        return

    db = SessionLocal()
    try:
        active_count = (
            db.query(OutboundTask)
            .filter(OutboundTask.status.in_(["starting", "started", "in_progress"]))
            .count()
        )
        free_slots = OUTBOUND_MAX_CONCURRENT_CALLS - active_count
        if free_slots <= 0:
            return

        now = datetime.utcnow()
        tasks = (
            db.query(OutboundTask)
            .join(Campaign, Campaign.id == OutboundTask.campaign_id)
            .filter(Campaign.status == "active")
            .filter(or_(Campaign.next_call_after.is_(None), Campaign.next_call_after <= now))
            .filter(OutboundTask.status.in_(["pending", "scheduled"]))
            .filter(OutboundTask.scheduled_at <= now)
            .order_by(OutboundTask.priority.asc(), OutboundTask.scheduled_at.asc(), OutboundTask.id.asc())
            .limit(free_slots)
            .all()
        )
        task_ids = [task.id for task in tasks]
    finally:
        db.close()

    for task_id in task_ids:
        await start_outbound_task(task_id)


async def create_campaign_from_rows(
    db: Session,
    rows: list[dict[str, Any]],
    source_filename: str,
    admin_chat_id: str,
    campaign_context: str = "",
) -> Campaign:
    campaign = Campaign(
        name=f"{Path(source_filename).stem} {datetime.utcnow().strftime('%Y-%m-%d %H:%M')}",
        status="paused",
        source_filename=source_filename,
        prompt_context=campaign_context,
        default_max_attempts=1,
        call_delay_seconds=OUTBOUND_DEFAULT_CALL_DELAY_SECONDS,
        created_by_chat_id=admin_chat_id,
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
    )
    db.add(campaign)
    db.flush()

    for idx, row in enumerate(rows, start=1):
        max_attempts = safe_int(row.get("max_attempts")) or campaign.default_max_attempts or 1
        scheduled_at = parse_datetime_maybe(row.get("call_after")) or datetime.utcnow()
        contact = OutboundContact(
            campaign_id=campaign.id,
            phone=row["phone"],
            name=safe_text(row.get("name")).strip() or None,
            company=safe_text(row.get("company")).strip() or None,
            city=safe_text(row.get("city")).strip() or None,
            source=safe_text(row.get("source")).strip() or None,
            task=safe_text(row.get("task")).strip() or None,
            context=safe_text(row.get("context")).strip() or None,
            preferred_time=safe_text(row.get("preferred_time")).strip() or None,
            timezone=safe_text(row.get("timezone")).strip() or None,
            raw_row_json=to_json_text(row.get("_raw")),
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow(),
        )
        db.add(contact)
        db.flush()

        db.add(
            OutboundTask(
                campaign_id=campaign.id,
                contact_id=contact.id,
                phone=contact.phone,
                status="pending",
                priority=100 + idx,
                scheduled_at=scheduled_at,
                attempt_count=0,
                max_attempts=max_attempts,
                created_at=datetime.utcnow(),
                updated_at=datetime.utcnow(),
            )
        )

    db.commit()
    db.refresh(campaign)
    return campaign


def web_campaign_payload(db: Session, campaign: Campaign) -> dict[str, Any]:
    return {**campaign.as_dict(), "stats": get_campaign_stats(db, campaign.id)}


def web_task_payload(task: Optional[OutboundTask]) -> Optional[dict[str, Any]]:
    return task.as_dict() if task else None


def web_contact_payload(contact: OutboundContact, task: Optional[OutboundTask] = None) -> dict[str, Any]:
    raw_row = json_dict(contact.raw_row_json)
    normalized = raw_row.get("normalized") if isinstance(raw_row.get("normalized"), dict) else {}
    return {
        **contact.as_dict(),
        "attendance_status": normalized.get("attendance_status", ""),
        "activity_type": normalized.get("activity_type", ""),
        "is_decision_maker": normalized.get("is_decision_maker", ""),
        "average_check": normalized.get("average_check", ""),
        "traffic_source": normalized.get("traffic_source", ""),
        "bot_impression": normalized.get("bot_impression", ""),
        "task": web_task_payload(task),
    }


def web_call_payload(call: Call) -> dict[str, Any]:
    payload = call.as_dict()
    payload["recording_download_url"] = f"/web/recordings/{call.voximplant_session_id}" if call.local_recording_path else None
    return payload


def web_campaign_detail(db: Session, campaign: Campaign) -> dict[str, Any]:
    contacts = (
        db.query(OutboundContact)
        .filter(OutboundContact.campaign_id == campaign.id)
        .order_by(OutboundContact.id.asc())
        .all()
    )
    tasks = (
        db.query(OutboundTask)
        .filter(OutboundTask.campaign_id == campaign.id)
        .order_by(OutboundTask.id.asc())
        .all()
    )
    task_by_contact_id = {task.contact_id: task for task in tasks}
    calls = (
        db.query(Call)
        .filter(Call.campaign_id == campaign.id)
        .order_by(Call.updated_at.desc(), Call.id.desc())
        .limit(100)
        .all()
    )
    events = (
        db.query(OutboundEvent)
        .filter(OutboundEvent.campaign_id == campaign.id)
        .order_by(OutboundEvent.created_at.desc(), OutboundEvent.id.desc())
        .limit(100)
        .all()
    )
    return {
        "campaign": web_campaign_payload(db, campaign),
        "contacts": [web_contact_payload(contact, task_by_contact_id.get(contact.id)) for contact in contacts],
        "tasks": [task.as_dict() for task in tasks],
        "calls": [web_call_payload(call) for call in calls],
        "events": [event.as_dict() for event in events],
    }


def reset_campaign_tasks(db: Session, campaign: Campaign):
    now = datetime.utcnow()
    tasks = db.query(OutboundTask).filter(OutboundTask.campaign_id == campaign.id).all()
    for task in tasks:
        task.status = "pending"
        task.scheduled_at = now
        task.started_at = None
        task.finished_at = None
        task.last_attempt_at = None
        task.next_attempt_at = None
        task.attempt_count = 0
        task.voximplant_session_id = None
        task.telegram_chat_id = None
        task.telegram_message_id = None
        task.telegram_message_refs_json = None
        task.start_result_json = None
        task.last_status = None
        task.last_status_message = None
        task.last_status_at = None
        task.last_status_payload_json = None
        task.result_status = None
        task.result_summary = None
        task.last_error = None
        task.updated_at = now

    campaign.status = "paused"
    campaign.next_call_after = None
    campaign.paused_at = now
    campaign.updated_at = now


def parse_uploaded_rows(filename: str, content: bytes) -> tuple[list[dict[str, Any]], str]:
    suffix = Path(filename or "").suffix.lower()
    if suffix not in {".csv", ".tsv", ".xlsx"}:
        raise HTTPException(status_code=400, detail="unsupported file format")
    raw_rows = read_xlsx_rows(content) if suffix == ".xlsx" else read_csv_rows(content, suffix)
    return normalize_import_rows(raw_rows)


def admin_only(handler):
    async def wrapper(message: Message):
        if not is_admin_chat(message.chat.id):
            await message.answer("<b>Доступ закрыт.</b>")
            return
        await handler(message)

    return wrapper


async def tg_help(message: Message):
    text = (
        "<b>Управление обзвоном</b>\n\n"
        "<b>/help</b> — справка\n"
        "<b>/upload</b> или <b>/template</b> — получить шаблон таблицы\n"
        "<b>/campaigns</b> — список кампаний\n"
        "<b>/run ID</b> — запустить кампанию\n"
        "<b>/pause ID</b> или <b>/stop ID</b> — поставить кампанию на паузу\n"
        "<b>/rerun ID</b> — подготовить повторный запуск кампании\n"
        "<b>/delay ID 30s</b> — изменить паузу между контактами\n"
        "<b>/status</b> — состояние сервера и очереди\n"
        "<b>/calls</b> — последние звонки\n\n"
        "Для загрузки базы отправьте файл <b>XLSX</b>, <b>CSV</b> или <b>TSV</b>. "
        "После загрузки кампания создается на паузе."
    )
    await message.answer(text)


async def configure_telegram_commands() -> tuple[str, Optional[str]]:
    if bot is None:
        return "skipped_no_bot", None
    try:
        await bot.set_my_commands(TELEGRAM_BOT_COMMANDS)
        logger.info("Telegram bot commands configured")
        return "configured", None
    except Exception as exc:  # noqa: BLE001
        logger.warning("Telegram set_my_commands failed: %s", str(exc))
        return "error", str(exc)


async def tg_setcommands(message: Message):
    status, error_text = await configure_telegram_commands()
    if status == "configured":
        await message.answer("<b>Меню команд обновлено.</b>")
    else:
        await message.answer(f"<b>Не удалось обновить команды.</b>\n{html_text(error_text or status)}")


async def tg_status(message: Message):
    db = SessionLocal()
    try:
        campaigns = db.query(Campaign).count()
        contacts = db.query(OutboundContact).count()
        tasks_total = db.query(OutboundTask).count()
        pending = db.query(OutboundTask).filter(OutboundTask.status.in_(["pending", "scheduled"])).count()
        active = db.query(OutboundTask).filter(OutboundTask.status.in_(["starting", "started", "in_progress"])).count()
        completed = db.query(OutboundTask).filter(OutboundTask.status == "completed").count()
        failed = db.query(OutboundTask).filter(OutboundTask.status == "failed").count()
    finally:
        db.close()

    await message.answer(
        "\n".join(
            [
                "<b>Статус системы</b>",
                "",
                format_stat_line("Кампаний", campaigns),
                format_stat_line("Контактов", contacts),
                format_stat_line("Задач всего", tasks_total),
                format_stat_line("В очереди", pending),
                format_stat_line("В работе", active),
                format_stat_line("Завершено", completed),
                format_stat_line("Ошибок", failed),
                "",
                format_stat_line("Обработчик очереди", "включен" if OUTBOUND_WORKER_ENABLED else "выключен"),
                format_stat_line("ID правила Voximplant", VOXIMPLANT_RULE_ID or "не задан"),
                format_stat_line("Пауза по умолчанию", format_delay(OUTBOUND_DEFAULT_CALL_DELAY_SECONDS)),
            ]
        )
    )


async def tg_campaigns(message: Message):
    db = SessionLocal()
    try:
        campaigns = db.query(Campaign).order_by(Campaign.id.desc()).limit(10).all()
        if not campaigns:
            await message.answer(
                "<b>Кампаний пока нет.</b>\n\n"
                "Отправьте файл <b>XLSX</b>, <b>CSV</b> или <b>TSV</b> с колонкой телефона."
            )
            return
        lines = ["<b>Кампании</b>"]
        for campaign in campaigns:
            stats = get_campaign_stats(db, campaign.id)
            lines.append("")
            lines.append(f"<b>Кампания #{campaign.id}</b>")
            lines.append(format_stat_line("Название", campaign.name))
            lines.append(format_stat_line("Статус", status_ru(campaign.status)))
            lines.append(format_stat_line("Контактов", stats.get("total", 0)))
            lines.append(format_stat_line("Ожидают", stats.get("pending", 0) + stats.get("scheduled", 0)))
            lines.append(
                format_stat_line(
                    "В работе",
                    stats.get("started", 0) + stats.get("starting", 0) + stats.get("in_progress", 0),
                )
            )
            lines.append(format_stat_line("Завершено", stats.get("completed", 0)))
            lines.append(format_stat_line("Ошибок", stats.get("failed", 0)))
            lines.append(format_stat_line("Пауза", format_delay(campaign.call_delay_seconds)))
            if campaign.next_call_after and campaign.next_call_after > datetime.utcnow():
                lines.append(format_stat_line("Следующий старт после", campaign.next_call_after.strftime("%H:%M:%S")))
    finally:
        db.close()
    await message.answer("\n".join(lines))


async def tg_run_or_pause(message: Message, status: str):
    parts = safe_text(message.text).split()
    if len(parts) < 2 or not parts[1].isdigit():
        await message.answer("<b>Укажите ID кампании.</b>\n\nПример: <code>/run 1</code>")
        return
    campaign_id = int(parts[1])
    db = SessionLocal()
    try:
        campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
        if not campaign:
            await message.answer(f"<b>Кампания #{campaign_id} не найдена.</b>")
            return
        campaign.status = status
        campaign.updated_at = datetime.utcnow()
        if status == "active":
            campaign.started_at = datetime.utcnow()
            campaign.next_call_after = None
        else:
            campaign.paused_at = datetime.utcnow()
        db.commit()
        stats = get_campaign_stats(db, campaign.id)
    finally:
        db.close()
    action = "запущена" if status == "active" else "остановлена"
    await message.answer(
        "\n".join(
            [
                f"<b>Кампания #{campaign_id} {action}.</b>",
                "",
                format_stat_line("Задач всего", stats.get("total", 0)),
                format_stat_line("Ожидают", stats.get("pending", 0) + stats.get("scheduled", 0)),
                format_stat_line("В работе", stats.get("started", 0) + stats.get("starting", 0) + stats.get("in_progress", 0)),
                format_stat_line("Завершено", stats.get("completed", 0)),
                format_stat_line("Ошибок", stats.get("failed", 0)),
            ]
        )
    )


async def tg_rerun(message: Message):
    parts = safe_text(message.text).split()
    if len(parts) < 2 or not parts[1].isdigit():
        await message.answer("<b>Укажите ID кампании.</b>\n\nПример: <code>/rerun 1</code>")
        return

    campaign_id = int(parts[1])
    db = SessionLocal()
    try:
        campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
        if not campaign:
            await message.answer(f"<b>Кампания #{campaign_id} не найдена.</b>")
            return

        now = datetime.utcnow()
        tasks = db.query(OutboundTask).filter(OutboundTask.campaign_id == campaign_id).all()
        for task in tasks:
            task.status = "pending"
            task.scheduled_at = now
            task.started_at = None
            task.finished_at = None
            task.last_attempt_at = None
            task.next_attempt_at = None
            task.attempt_count = 0
            task.voximplant_session_id = None
            task.telegram_chat_id = None
            task.telegram_message_id = None
            task.telegram_message_refs_json = None
            task.start_result_json = None
            task.last_status = None
            task.last_status_message = None
            task.last_status_at = None
            task.last_status_payload_json = None
            task.result_status = None
            task.result_summary = None
            task.last_error = None
            task.updated_at = now

        campaign.status = "paused"
        campaign.next_call_after = None
        campaign.paused_at = now
        campaign.updated_at = now
        db.commit()
        stats = get_campaign_stats(db, campaign.id)
    finally:
        db.close()

    await message.answer(
        "\n".join(
            [
                f"<b>Кампания #{campaign_id} подготовлена к повторному запуску.</b>",
                "",
                format_stat_line("Статус", "на паузе"),
                format_stat_line("Задач сброшено", stats.get("pending", 0) + stats.get("scheduled", 0)),
                "",
                f"Запуск: <code>/run {campaign_id}</code>",
            ]
        )
    )


async def tg_delay(message: Message):
    parts = safe_text(message.text).split(maxsplit=2)
    if len(parts) < 3 or not parts[1].isdigit():
        await message.answer("<b>Укажите ID кампании и паузу.</b>\n\nПримеры: <code>/delay 1 30s</code> или <code>/delay 1 2m</code>")
        return

    delay_seconds = parse_delay_seconds(parts[2])
    if delay_seconds is None:
        await message.answer("<b>Не понял паузу.</b>\n\nПримеры: <code>30s</code>, <code>45 сек</code>, <code>2m</code>, <code>2 мин</code>.")
        return

    campaign_id = int(parts[1])
    db = SessionLocal()
    try:
        campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
        if not campaign:
            await message.answer(f"<b>Кампания #{campaign_id} не найдена.</b>")
            return
        campaign.call_delay_seconds = delay_seconds
        campaign.updated_at = datetime.utcnow()
        db.commit()
    finally:
        db.close()

    await message.answer(
        "\n".join(
            [
                f"<b>Пауза обновлена.</b>",
                "",
                format_stat_line("Кампания", f"#{campaign_id}"),
                format_stat_line("Пауза между контактами", format_delay(delay_seconds)),
            ]
        )
    )


async def tg_template(message: Message):
    template_path = ensure_forum_amix_template()
    caption = (
        "<b>Шаблон базы обзвона</b>\n\n"
        "<b>Форматы:</b> XLSX, CSV, TSV.\n"
        "Для Amix используем этот шаблон: до столбца G — данные от организатора, после G — поля менеджера и результат прозвона.\n\n"
        "<b>Ключевые колонки:</b>\n"
        "<b>E:</b> Контактныйтелефон\n"
        "<b>G:</b> Пришёл — по нему выбирается начало скрипта\n"
        "<b>B/C/F:</b> имя, фамилия, компания\n"
        "<b>I-P:</b> поля, которые бот может учитывать и потом заполнять через Google Sheets.\n\n"
        "Отправьте заполненный файл сюда. Кампания создастся на паузе."
    )
    await message.answer_document(FSInputFile(str(template_path)), caption=caption)


async def tg_calls(message: Message):
    db = SessionLocal()
    try:
        calls = db.query(Call).order_by(Call.id.desc()).limit(5).all()
        if not calls:
            await message.answer("<b>Звонков пока нет.</b>")
            return
        lines = ["<b>Последние звонки</b>"]
        for call in calls:
            lines.append("")
            lines.append(f"<b>Звонок #{call.id}</b>")
            lines.append(format_stat_line("Номер", call.client_phone or call.caller_phone or "не указан"))
            lines.append(format_stat_line("Статус", status_ru(call.status)))
            if call.client_name:
                lines.append(format_stat_line("Имя", call.client_name))
            if call.duration is not None:
                lines.append(format_stat_line("Длительность", f"{call.duration} сек"))
            if call.outcome:
                lines.append(format_stat_line("Итог", call.outcome))
            lines.append(format_stat_line("Кратко", call.summary or "итоги пока не сохранены"))
    finally:
        db.close()
    await message.answer("\n".join(lines))


async def tg_document(message: Message):
    document = message.document
    if not document:
        return
    suffix = Path(document.file_name or "").suffix.lower()
    if suffix not in {".csv", ".tsv", ".xlsx"}:
        await message.answer("<b>Формат не поддерживается.</b>\n\nМожно загрузить файл <b>XLSX</b>, <b>CSV</b> или <b>TSV</b>.")
        return

    file = await bot.get_file(document.file_id)  # type: ignore[union-attr]
    buffer = io.BytesIO()
    await bot.download_file(file.file_path, destination=buffer)  # type: ignore[union-attr]
    content = buffer.getvalue()

    saved_path = IMPORTS_DIR / f"{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{document.file_name}"
    saved_path.write_bytes(content)

    try:
        raw_rows = read_xlsx_rows(content) if suffix == ".xlsx" else read_csv_rows(content, suffix)
        rows, campaign_context = normalize_import_rows(raw_rows)
    except Exception as exc:  # noqa: BLE001
        logger.exception("Failed to parse uploaded file")
        await message.answer(f"<b>Не удалось разобрать файл.</b>\n\n{html_text(exc)}")
        return

    if not rows:
        await message.answer(
            "<b>В файле не нашел контактов.</b>\n\n"
            "Проверьте колонку <b>phone</b>/<b>телефон</b> или колонку <b>E</b> "
            "в шаблоне Amix, если файл загружен без строки заголовков."
        )
        return

    db = SessionLocal()
    try:
        campaign = await create_campaign_from_rows(
            db,
            rows=rows,
            source_filename=document.file_name or saved_path.name,
            admin_chat_id=safe_text(message.chat.id),
            campaign_context=campaign_context,
        )
        stats = get_campaign_stats(db, campaign.id)
    finally:
        db.close()

    await message.answer(
        "\n".join(
            [
                f"<b>База загружена.</b>",
                "",
                format_stat_line("Кампания", f"#{campaign.id}"),
                format_stat_line("Контактов", stats.get("total", 0)),
                format_stat_line("Статус", status_ru(campaign.status)),
                format_stat_line("Пауза между контактами", format_delay(campaign.call_delay_seconds)),
                "",
                f"Запуск: <code>/run {campaign.id}</code>",
                f"Изменить паузу: <code>/delay {campaign.id} 30s</code>",
            ]
        )
    )


def setup_telegram_dispatcher():
    global dispatcher
    if bot is None:
        return
    dispatcher = Dispatcher()
    dispatcher.message.register(admin_only(tg_help), Command("start", "help"))
    dispatcher.message.register(admin_only(tg_status), Command("status"))
    dispatcher.message.register(admin_only(tg_campaigns), Command("campaigns"))
    dispatcher.message.register(admin_only(lambda message: tg_run_or_pause(message, "active")), Command("run"))
    dispatcher.message.register(admin_only(tg_rerun), Command("rerun"))
    dispatcher.message.register(admin_only(lambda message: tg_run_or_pause(message, "paused")), Command("pause"))
    dispatcher.message.register(admin_only(lambda message: tg_run_or_pause(message, "paused")), Command("stop"))
    dispatcher.message.register(admin_only(tg_delay), Command("delay", "interval"))
    dispatcher.message.register(admin_only(tg_template), Command("template", "upload"))
    dispatcher.message.register(admin_only(tg_calls), Command("calls"))
    dispatcher.message.register(admin_only(tg_setcommands), Command("setcommands"))
    dispatcher.message.register(admin_only(tg_document), F.document)


@asynccontextmanager
async def lifespan(app: FastAPI):
    global telegram_polling_task
    logger.info("Artem Fitisov backend starting")
    ensure_forum_amix_template()
    scheduler.add_job(cleanup_old_recordings, "interval", hours=CLEANUP_INTERVAL_HOURS)
    scheduler.add_job(process_outbound_queue, "interval", seconds=OUTBOUND_WORKER_INTERVAL_SECONDS)
    scheduler.start()
    setup_telegram_dispatcher()
    if dispatcher is not None and bot is not None:
        await configure_telegram_commands()
        telegram_polling_task = asyncio.create_task(dispatcher.start_polling(bot))
        logger.info("Telegram bot polling started")
    yield
    logger.info("Artem Fitisov backend stopping")
    if telegram_polling_task:
        telegram_polling_task.cancel()
        try:
            await telegram_polling_task
        except asyncio.CancelledError:
            pass
    scheduler.shutdown()
    if bot is not None:
        await bot.session.close()


app = FastAPI(title="Artem Fitisov Backend", lifespan=lifespan)

if WEB_CORS_ORIGINS:
    app.add_middleware(
        CORSMiddleware,
        allow_origins=WEB_CORS_ORIGINS,
        allow_credentials=True,
        allow_methods=["*"],
        allow_headers=["*"],
    )


@app.get("/")
def read_root():
    return {"service": "artem_fitisov_backend", "status": "ok", "time_utc": datetime.utcnow().isoformat()}


@app.get("/healthz")
def healthcheck():
    return {"ok": True, "time_utc": datetime.utcnow().isoformat()}


@app.post("/web/auth/login")
def web_login(payload: WebLoginPayload, response: Response):
    if not WEB_ADMIN_PASSWORD:
        raise HTTPException(status_code=503, detail="web auth is not configured")
    username_ok = hmac.compare_digest(payload.username.strip(), WEB_ADMIN_USERNAME)
    password_ok = hmac.compare_digest(payload.password, WEB_ADMIN_PASSWORD)
    if not username_ok or not password_ok:
        raise HTTPException(status_code=401, detail="invalid credentials")

    set_web_session_cookie(response, WEB_ADMIN_USERNAME)
    return {"ok": True, "username": WEB_ADMIN_USERNAME}


@app.post("/web/auth/logout")
def web_logout(response: Response):
    clear_web_session_cookie(response)
    return {"ok": True}


@app.get("/web/auth/me")
def web_me(username: str = Depends(require_web_user)):
    return {"username": username}


@app.get("/web/dashboard")
def web_dashboard(
    username: str = Depends(require_web_user),
    db: Session = Depends(get_db),
):
    del username
    task_pending = db.query(OutboundTask).filter(OutboundTask.status.in_(["pending", "scheduled"])).count()
    task_active = db.query(OutboundTask).filter(OutboundTask.status.in_(["starting", "started", "in_progress"])).count()
    task_completed = db.query(OutboundTask).filter(OutboundTask.status == "completed").count()
    task_failed = db.query(OutboundTask).filter(OutboundTask.status == "failed").count()
    campaigns_total = db.query(Campaign).count()
    recent_calls = db.query(Call).order_by(Call.updated_at.desc(), Call.id.desc()).limit(8).all()
    recent_campaigns = db.query(Campaign).order_by(Campaign.updated_at.desc(), Campaign.id.desc()).limit(8).all()
    return {
        "campaigns": {
            "total": campaigns_total,
            "active": db.query(Campaign).filter(Campaign.status == "active").count(),
            "paused": db.query(Campaign).filter(Campaign.status == "paused").count(),
            "recent": [web_campaign_payload(db, campaign) for campaign in recent_campaigns],
        },
        "contacts": {"total": db.query(OutboundContact).count()},
        "tasks": {
            "total": db.query(OutboundTask).count(),
            "pending": task_pending,
            "active": task_active,
            "completed": task_completed,
            "failed": task_failed,
        },
        "calls": {
            "total": db.query(Call).count(),
            "recent": [web_call_payload(call) for call in recent_calls],
        },
        "worker": {
            "enabled": OUTBOUND_WORKER_ENABLED,
            "rule_id": VOXIMPLANT_RULE_ID or None,
            "public_backend_url": PUBLIC_BACKEND_URL or None,
            "queue_interval_seconds": OUTBOUND_WORKER_INTERVAL_SECONDS,
            "max_concurrent_calls": OUTBOUND_MAX_CONCURRENT_CALLS,
        },
    }


@app.get("/web/campaigns")
def web_list_campaigns(
    username: str = Depends(require_web_user),
    db: Session = Depends(get_db),
):
    del username
    campaigns = db.query(Campaign).order_by(Campaign.id.desc()).limit(200).all()
    return [web_campaign_payload(db, campaign) for campaign in campaigns]


@app.get("/web/campaigns/{campaign_id}")
def web_get_campaign(
    campaign_id: int,
    username: str = Depends(require_web_user),
    db: Session = Depends(get_db),
):
    del username
    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="campaign not found")
    return web_campaign_detail(db, campaign)


@app.post("/web/imports")
async def web_upload_import(
    file: UploadFile = File(...),
    username: str = Depends(require_web_user),
    db: Session = Depends(get_db),
):
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="empty file")

    rows, campaign_context = parse_uploaded_rows(file.filename or "upload", content)
    if not rows:
        raise HTTPException(status_code=400, detail="no contacts found")

    saved_name = f"{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{Path(file.filename or 'upload').name}"
    (IMPORTS_DIR / saved_name).write_bytes(content)
    campaign = await create_campaign_from_rows(
        db,
        rows=rows,
        source_filename=file.filename or saved_name,
        admin_chat_id=f"web:{username}",
        campaign_context=campaign_context,
    )
    return {
        "campaign": campaign.as_dict(),
        "stats": get_campaign_stats(db, campaign.id),
        "preview": rows[:20],
    }


@app.get("/web/template")
def web_download_template(username: str = Depends(require_web_user)):
    del username
    template_path = ensure_forum_amix_template()
    return FileResponse(
        path=str(template_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=template_path.name,
    )


@app.post("/web/campaigns/{campaign_id}/run")
def web_run_campaign(
    campaign_id: int,
    username: str = Depends(require_web_user),
    db: Session = Depends(get_db),
):
    del username
    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="campaign not found")
    campaign.status = "active"
    campaign.started_at = datetime.utcnow()
    campaign.next_call_after = None
    campaign.updated_at = datetime.utcnow()
    db.commit()
    return web_campaign_payload(db, campaign)


@app.post("/web/campaigns/{campaign_id}/pause")
def web_pause_campaign(
    campaign_id: int,
    username: str = Depends(require_web_user),
    db: Session = Depends(get_db),
):
    del username
    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="campaign not found")
    campaign.status = "paused"
    campaign.paused_at = datetime.utcnow()
    campaign.updated_at = datetime.utcnow()
    db.commit()
    return web_campaign_payload(db, campaign)


@app.post("/web/campaigns/{campaign_id}/rerun")
def web_rerun_campaign(
    campaign_id: int,
    username: str = Depends(require_web_user),
    db: Session = Depends(get_db),
):
    del username
    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="campaign not found")
    reset_campaign_tasks(db, campaign)
    db.commit()
    return web_campaign_payload(db, campaign)


@app.post("/web/campaigns/{campaign_id}/delay")
def web_update_campaign_delay(
    campaign_id: int,
    payload: WebDelayPayload,
    username: str = Depends(require_web_user),
    db: Session = Depends(get_db),
):
    del username
    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="campaign not found")
    campaign.call_delay_seconds = payload.delay_seconds
    campaign.updated_at = datetime.utcnow()
    db.commit()
    return web_campaign_payload(db, campaign)


@app.patch("/web/contacts/{contact_id}")
def web_update_contact(
    contact_id: int,
    payload: WebContactUpdatePayload,
    username: str = Depends(require_web_user),
    db: Session = Depends(get_db),
):
    del username
    contact = db.query(OutboundContact).filter(OutboundContact.id == contact_id).first()
    if not contact:
        raise HTTPException(status_code=404, detail="contact not found")

    changes = payload.model_dump(exclude_unset=True)
    if "phone" in changes:
        normalized_phone = normalize_phone(changes["phone"])
        if not normalized_phone:
            raise HTTPException(status_code=400, detail="phone is empty")
        contact.phone = normalized_phone
        db.query(OutboundTask).filter(OutboundTask.contact_id == contact.id).update({"phone": normalized_phone})

    for field_name in ("name", "company", "city", "source", "task", "context", "preferred_time", "timezone"):
        if field_name in changes:
            setattr(contact, field_name, safe_text(changes[field_name]).strip() or None)

    contact.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(contact)
    task = (
        db.query(OutboundTask)
        .filter(OutboundTask.contact_id == contact.id)
        .order_by(OutboundTask.id.desc())
        .first()
    )
    return web_contact_payload(contact, task)


@app.get("/web/calls")
def web_list_calls(
    username: str = Depends(require_web_user),
    limit: int = Query(default=50, ge=1, le=500),
    db: Session = Depends(get_db),
):
    del username
    calls = db.query(Call).order_by(Call.updated_at.desc(), Call.id.desc()).limit(limit).all()
    return [web_call_payload(call) for call in calls]


@app.get("/web/call-history")
def web_call_history(
    phone: str = Query(..., min_length=3),
    username: str = Depends(require_web_user),
    limit: int = Query(default=100, ge=1, le=300),
    db: Session = Depends(get_db),
):
    del username
    normalized_phone = normalize_phone(phone)
    if not normalized_phone:
        raise HTTPException(status_code=400, detail="phone is empty")

    calls = (
        db.query(Call)
        .filter(or_(Call.client_phone == normalized_phone, Call.caller_phone == normalized_phone))
        .order_by(Call.updated_at.desc(), Call.id.desc())
        .limit(limit)
        .all()
    )
    return [web_call_payload(call) for call in calls]


@app.get("/web/calls/{session_id}")
def web_get_call(
    session_id: str = ApiPath(..., min_length=3),
    username: str = Depends(require_web_user),
    db: Session = Depends(get_db),
):
    del username
    db_call = db.query(Call).filter(Call.voximplant_session_id == session_id).first()
    if not db_call:
        raise HTTPException(status_code=404, detail="call not found")
    return web_call_payload(db_call)


@app.get("/web/recordings/{session_id}")
def web_get_recording(
    session_id: str = ApiPath(..., min_length=3),
    username: str = Depends(require_web_user),
    db: Session = Depends(get_db),
):
    del username
    db_call = db.query(Call).filter(Call.voximplant_session_id == session_id).first()
    if not db_call or not db_call.local_recording_path:
        raise HTTPException(status_code=404, detail="recording not found")
    file_path = Path(db_call.local_recording_path)
    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(status_code=404, detail="recording file not found")
    return FileResponse(path=str(file_path), filename=file_path.name)


@app.get("/calls")
def list_calls(
    secret: str = Query(default=""),
    limit: int = Query(default=50, ge=1, le=500),
    db: Session = Depends(get_db),
):
    if not BACKEND_WEBHOOK_SECRET:
        raise HTTPException(status_code=503, detail="admin api disabled")
    if secret != BACKEND_WEBHOOK_SECRET:
        raise HTTPException(status_code=403, detail="invalid secret")

    calls = db.query(Call).order_by(Call.started_at.desc()).limit(limit).all()
    return [call.as_dict() for call in calls]


@app.get("/calls/{session_id}")
def get_call(
    session_id: str = ApiPath(..., min_length=3),
    secret: str = Query(default=""),
    db: Session = Depends(get_db),
):
    if not BACKEND_WEBHOOK_SECRET:
        raise HTTPException(status_code=503, detail="admin api disabled")
    if secret != BACKEND_WEBHOOK_SECRET:
        raise HTTPException(status_code=403, detail="invalid secret")

    db_call = db.query(Call).filter(Call.voximplant_session_id == session_id).first()
    if not db_call:
        raise HTTPException(status_code=404, detail="call not found")
    return db_call.as_dict()


@app.get("/outbound/campaigns")
def list_campaigns(
    secret: str = Query(default=""),
    db: Session = Depends(get_db),
):
    if not BACKEND_WEBHOOK_SECRET:
        raise HTTPException(status_code=503, detail="admin api disabled")
    if secret != BACKEND_WEBHOOK_SECRET:
        raise HTTPException(status_code=403, detail="invalid secret")

    campaigns = db.query(Campaign).order_by(Campaign.id.desc()).limit(100).all()
    return [{**campaign.as_dict(), "stats": get_campaign_stats(db, campaign.id)} for campaign in campaigns]


@app.post("/outbound/campaigns/{campaign_id}/run")
def run_campaign(
    campaign_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    require_webhook_secret(request)
    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="campaign not found")
    campaign.status = "active"
    campaign.started_at = datetime.utcnow()
    campaign.updated_at = datetime.utcnow()
    db.commit()
    return {"status": "active", "campaign_id": campaign_id, "stats": get_campaign_stats(db, campaign_id)}


@app.post("/outbound/campaigns/{campaign_id}/pause")
def pause_campaign(
    campaign_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    require_webhook_secret(request)
    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="campaign not found")
    campaign.status = "paused"
    campaign.paused_at = datetime.utcnow()
    campaign.updated_at = datetime.utcnow()
    db.commit()
    return {"status": "paused", "campaign_id": campaign_id, "stats": get_campaign_stats(db, campaign_id)}


@app.get("/outbound/tasks/{task_id}/scenario-context")
def get_scenario_context(
    task_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    require_webhook_secret(request)

    task = db.query(OutboundTask).filter(OutboundTask.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="task not found")
    contact = db.query(OutboundContact).filter(OutboundContact.id == task.contact_id).first()
    campaign = db.query(Campaign).filter(Campaign.id == task.campaign_id).first()
    if not contact or not campaign:
        raise HTTPException(status_code=404, detail="context not found")

    return build_task_context(campaign, contact, task)


@app.get("/inbound/caller-context")
def get_inbound_caller_context(
    request: Request,
    phone: str = Query(..., min_length=3),
    db: Session = Depends(get_db),
):
    require_webhook_secret(request)

    normalized_phone = normalize_phone(phone)
    if not normalized_phone:
        raise HTTPException(status_code=400, detail="phone is empty")

    latest_contact = (
        db.query(OutboundContact)
        .filter(OutboundContact.phone == normalized_phone)
        .order_by(OutboundContact.updated_at.desc(), OutboundContact.id.desc())
        .first()
    )
    recent_tasks = (
        db.query(OutboundTask)
        .filter(OutboundTask.phone == normalized_phone)
        .order_by(OutboundTask.updated_at.desc(), OutboundTask.id.desc())
        .limit(5)
        .all()
    )
    recent_calls = (
        db.query(Call)
        .filter(or_(Call.client_phone == normalized_phone, Call.caller_phone == normalized_phone))
        .order_by(Call.updated_at.desc(), Call.id.desc())
        .limit(5)
        .all()
    )

    latest_task = recent_tasks[0] if recent_tasks else None
    context_contact = latest_contact
    context_campaign = None
    lead_context: dict[str, Any] = {}

    if latest_task:
        task_contact = db.query(OutboundContact).filter(OutboundContact.id == latest_task.contact_id).first()
        task_campaign = db.query(Campaign).filter(Campaign.id == latest_task.campaign_id).first()
        if task_contact and task_campaign:
            context_contact = task_contact
            context_campaign = task_campaign
            lead_context = build_task_context(task_campaign, task_contact, latest_task)

    if not lead_context and context_contact:
        context_campaign = db.query(Campaign).filter(Campaign.id == context_contact.campaign_id).first()
        raw_row = json_dict(context_contact.raw_row_json)
        normalized = raw_row.get("normalized") if isinstance(raw_row.get("normalized"), dict) else {}
        lead_context = {
            "phone": context_contact.phone,
            "client_name": context_contact.name or "",
            "last_name": normalized.get("last_name", ""),
            "email": normalized.get("email", ""),
            "company": context_contact.company or "",
            "city": context_contact.city or "",
            "source": context_contact.source or "",
            "attendance_status": normalized.get("attendance_status", ""),
            "activity_type": normalized.get("activity_type", ""),
            "is_decision_maker": normalized.get("is_decision_maker", ""),
            "average_check": normalized.get("average_check", ""),
            "traffic_source": normalized.get("traffic_source", ""),
            "bot_impression": normalized.get("bot_impression", ""),
            "task": context_contact.task or "",
            "context": context_contact.context or "",
            "preferred_time": context_contact.preferred_time or "",
            "timezone": context_contact.timezone or "",
            "campaign_context": context_campaign.prompt_context if context_campaign else "",
        }

    task_briefs: list[dict[str, Any]] = []
    for task in recent_tasks:
        contact = context_contact if context_contact and context_contact.id == task.contact_id else None
        if not contact:
            contact = db.query(OutboundContact).filter(OutboundContact.id == task.contact_id).first()
        campaign = context_campaign if context_campaign and context_campaign.id == task.campaign_id else None
        if not campaign:
            campaign = db.query(Campaign).filter(Campaign.id == task.campaign_id).first()
        task_briefs.append(task_context_brief(task, contact, campaign))

    call_briefs = [call_context_brief(db_call) for db_call in recent_calls]
    last_task = task_briefs[0] if task_briefs else None
    last_call = call_briefs[0] if call_briefs else None

    return {
        "phone": normalized_phone,
        "known": bool(lead_context or task_briefs or call_briefs),
        "lead_context": lead_context,
        "last_task": last_task,
        "last_call": last_call,
        "recent_tasks": task_briefs,
        "recent_calls": call_briefs,
        "context_text": build_inbound_context_text(normalized_phone, lead_context, last_task, last_call),
    }


@app.post("/webhook/voximplant/status")
async def voximplant_status(
    payload: VoximplantStatusPayload,
    request: Request,
    db: Session = Depends(get_db),
):
    require_webhook_secret(request)

    now = parse_iso_datetime(payload.timestamp_utc) or datetime.utcnow()
    stage = safe_text(payload.stage).strip() or "unknown"
    message = safe_text(payload.message).strip()
    payload_json = to_json_text(payload.model_dump(mode="json"))

    event = OutboundEvent(
        campaign_id=payload.campaign_id,
        task_id=payload.outbound_task_id,
        session_id=safe_text(payload.session_id).strip() or None,
        phone=normalize_phone(payload.phone),
        stage=stage,
        status=safe_text(payload.status).strip() or None,
        message=message or None,
        payload_json=payload_json,
        created_at=now,
    )
    db.add(event)

    task = None
    if payload.outbound_task_id:
        task = db.query(OutboundTask).filter(OutboundTask.id == payload.outbound_task_id).first()
    if not task and payload.session_id:
        task = db.query(OutboundTask).filter(OutboundTask.voximplant_session_id == payload.session_id).first()

    if task:
        task.last_status = stage
        task.last_status_message = message or None
        task.last_status_at = now
        task.last_status_payload_json = payload_json
        if payload.session_id and not task.voximplant_session_id:
            task.voximplant_session_id = payload.session_id
        task.updated_at = datetime.utcnow()
        if stage in {"dial_error", "empty_call_target"}:
            task.last_error = message or stage
        if stage == "call_connected" and task.status in {"starting", "started"}:
            task.status = "in_progress"

    db.commit()

    if task:
        asyncio.create_task(notify_outbound_task(task.id, status_label(stage, message or None)))

    return {"status": "success", "stage": stage, "event_id": event.id}


@app.post("/webhook/voximplant/call_started")
async def call_started(
    payload: CallStartedPayload,
    request: Request,
    db: Session = Depends(get_db),
):
    require_webhook_secret(request)

    db_call = get_or_create_call(db, payload.session_id)
    if payload.outbound_task_id:
        db_call.outbound_task_id = payload.outbound_task_id
    if payload.campaign_id:
        db_call.campaign_id = payload.campaign_id
    set_if_value(db_call, "project", payload.project)
    set_if_value(db_call, "script_name", payload.script_name)
    set_if_value(db_call, "caller_phone", payload.caller_phone)

    connected_at = parse_iso_datetime(payload.connected_at_utc)
    if connected_at:
        db_call.connected_at = connected_at
    if not db_call.started_at:
        db_call.started_at = datetime.utcnow()
    db_call.updated_at = datetime.utcnow()
    db_call.status = db_call.status or "started"

    if payload.outbound_task_id:
        task = db.query(OutboundTask).filter(OutboundTask.id == payload.outbound_task_id).first()
        if task:
            task.status = "in_progress"
            task.voximplant_session_id = payload.session_id
            task.started_at = db_call.connected_at or datetime.utcnow()
            task.updated_at = datetime.utcnow()
    db.commit()
    if payload.outbound_task_id:
        asyncio.create_task(notify_outbound_task(payload.outbound_task_id, "Соединение установлено"))

    return {"status": "success", "session_id": payload.session_id}


@app.post("/webhook/voximplant/call_finished")
async def call_finished(
    payload: CallFinishedPayload,
    request: Request,
    db: Session = Depends(get_db),
):
    require_webhook_secret(request)

    db_call = get_or_create_call(db, payload.session_id)
    if payload.duration is not None:
        db_call.duration = safe_int(payload.duration)
    set_if_value(db_call, "summary", payload.summary)
    db_call.finished_at = parse_iso_datetime(payload.finished_at_utc) or datetime.utcnow()
    db_call.updated_at = datetime.utcnow()
    db_call.status = payload.status or "finished"
    db.commit()

    return {"status": "success", "session_id": payload.session_id}


@app.post("/webhook/voximplant/recording_ready")
async def recording_ready(
    payload: RecordingReadyPayload,
    request: Request,
    db: Session = Depends(get_db),
):
    require_webhook_secret(request)

    db_call = get_or_create_call(db, payload.session_id)
    if payload.outbound_task_id:
        db_call.outbound_task_id = payload.outbound_task_id
    if payload.campaign_id:
        db_call.campaign_id = payload.campaign_id
    set_if_value(db_call, "project", payload.project)
    set_if_value(db_call, "script_name", payload.script_name)
    db_call.recording_url = payload.recording_url
    db_call.recording_status = payload.recording_status or "ready"
    set_if_value(db_call, "recording_error", payload.recording_error)
    db_call.updated_at = datetime.utcnow()
    db.commit()

    if not payload.outbound_task_id:
        asyncio.create_task(persist_recording_download(payload.session_id, payload.recording_url))
    return {"status": "success", "session_id": payload.session_id}


@app.post("/webhook/voximplant/report")
async def legacy_report(
    payload: CallReportPayload,
    request: Request,
):
    require_webhook_secret(request)

    if payload.report_type == "ADMIN_REPORT":
        status, error_text = await send_telegram_text(get_admin_chat_ids(), payload.text)
    elif payload.report_type == "SUMMARY_REPORT":
        status, error_text = await send_telegram_text(get_summary_chat_ids(), payload.text)
    else:
        raise HTTPException(status_code=400, detail="unknown report_type")

    return {"status": status, "error": error_text}


@app.post("/webhook/voximplant/google_sheets")
async def legacy_google_sheets(
    request: Request,
):
    require_webhook_secret(request)
    payload = await request.json()
    status, response_text = await send_to_google_sheets(payload)
    return {"status": status, "response": response_text}


@app.post("/webhook/voximplant/finalize")
async def finalize_call(
    payload: FinalizePayload,
    request: Request,
    db: Session = Depends(get_db),
):
    require_webhook_secret(request)

    db_call = get_or_create_call(db, payload.session_id)
    fill_call_from_finalize(db_call, payload)
    update_outbound_task_from_finalize(db, payload)
    db.commit()

    recording_download_status = None
    recording_download_error = None
    recording_url_for_download = payload.recording_url or db_call.recording_url
    if recording_url_for_download:
        if payload.outbound_task_id:
            recording_download_status = "scheduled"
            asyncio.create_task(
                persist_and_send_outbound_recording(
                    payload.outbound_task_id,
                    payload.session_id,
                    recording_url_for_download,
                )
            )
        else:
            recording_download_status = "scheduled"

    inbound_recording_reply_refs: list[dict[str, Any]] = []
    if is_diagnostic_finalize(payload):
        admin_status, admin_error = "skipped_diagnostic", None
        summary_status, summary_error = "skipped_diagnostic", None
    elif payload.outbound_task_id:
        admin_status, admin_error = "skipped_outbound_task", None
        summary_status, summary_error = "skipped_outbound_task", None
    else:
        admin_status, admin_error = "skipped_inbound_concise", None
        summary_status, summary_error, inbound_recording_reply_refs = await send_telegram_text_with_refs(
            get_summary_chat_ids(),
            render_inbound_summary_report(payload),
        )

    sheets_payload = payload.model_dump(mode="json")
    sheets_status, sheets_response = await send_to_google_sheets(sheets_payload)

    db_call.telegram_admin_status = admin_status
    db_call.telegram_summary_status = summary_status
    db_call.google_sheets_status = sheets_status
    db_call.google_sheets_response = sheets_response

    error_parts = [part for part in [admin_error, summary_error] if part]
    if sheets_status == "error" and sheets_response:
        error_parts.append(sheets_response)
    db_call.last_error = " | ".join(error_parts) if error_parts else None
    db_call.updated_at = datetime.utcnow()
    db.commit()

    recording_send_status = None
    recording_send_error = None
    if payload.outbound_task_id:
        await notify_outbound_task(payload.outbound_task_id, "Звонок завершен")
        recording_send_status = "scheduled" if recording_url_for_download else "skipped_no_recording_url"
    elif not is_diagnostic_finalize(payload):
        if recording_url_for_download:
            recording_send_status = "scheduled"
            asyncio.create_task(
                persist_and_send_inbound_recording(
                    payload.session_id,
                    recording_url_for_download,
                    inbound_recording_reply_refs,
                )
            )
        else:
            recording_send_status = "skipped_no_recording_url"

    return {
        "status": "success",
        "session_id": payload.session_id,
        "telegram_admin_status": admin_status,
        "telegram_summary_status": summary_status,
        "google_sheets_status": sheets_status,
        "recording_download_status": recording_download_status,
        "recording_download_error": recording_download_error,
        "recording_send_status": recording_send_status,
    }


if __name__ == "__main__":
    import uvicorn

    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
